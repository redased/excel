"""
Consolidation Dialog for PyQt6
Allows users to select multiple Excel files and configure consolidation
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QSpinBox, QComboBox,
    QGroupBox, QFormLayout, QFileDialog, QMessageBox, QProgressBar,
    QTextEdit, QCheckBox, QFrame, QDialogButtonBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from openpyxl import load_workbook

from core.consolidator import ConsolidationConfig, FileSource, ExcelConsolidator, ConsolidationResult


class ConsolidationWorker(QThread):
    """Worker thread for consolidation"""
    finished = pyqtSignal(object)
    progress = pyqtSignal(str)
    
    def __init__(self, consolidator, output_path):
        super().__init__()
        self.consolidator = consolidator
        self.output_path = output_path
    
    def run(self):
        self.progress.emit("Consolidation en cours...")
        result = self.consolidator.consolidate(self.output_path)
        self.finished.emit(result)


class AIAnalysisWorker(QThread):
    """Worker thread for AI analysis"""
    finished = pyqtSignal(object)
    
    def __init__(self, files, description, api_key):
        super().__init__()
        self.files = files
        self.description = description
        self.api_key = api_key
    
    def run(self):
        try:
            from core.consolidation_ai import analyze_for_consolidation
            result = analyze_for_consolidation(self.files, self.description, self.api_key)
            self.finished.emit(result)
        except Exception as e:
            self.finished.emit(e)


class SheetSelectionDialog(QDialog):
    """Dialog to select sheets from a file"""
    def __init__(self, sheets, selected=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sélectionner les feuilles")
        self.setModal(True)
        self.selected_sheets = selected or []
        
        layout = QVBoxLayout(self)
        
        scroll = QFrame()
        scroll_layout = QVBoxLayout(scroll)
        
        self.checkboxes = []
        for sheet in sheets:
            cb = QCheckBox(sheet)
            if sheet in self.selected_sheets:
                cb.setChecked(True)
            elif not self.selected_sheets: # Default select all if none selected
                cb.setChecked(True)
            self.checkboxes.append(cb)
            scroll_layout.addWidget(cb)
            
        layout.addWidget(scroll)
        
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
        
    def get_selected(self):
        return [cb.text() for cb in self.checkboxes if cb.isChecked()]


class ConsolidationDialog(QDialog):
    """Dialog for Excel file consolidation"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.consolidator = ExcelConsolidator()
        self.files = []
        self.worker = None
        self._setup_ui()
    
    def _setup_ui(self):
        self.setWindowTitle("📊 Consolidation de Fichiers Excel")
        self.setMinimumSize(1000, 700)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QLabel("📊 Consolidation de Fichiers Excel")
        header.setStyleSheet("font-size: 20px; font-weight: bold; color: #a78bfa; padding: 10px;")
        layout.addWidget(header)
        
        desc = QLabel(
            "Combinez plusieurs fichiers Excel et sélectionnez les feuilles à inclure.\n"
            "Vous pouvez utiliser l'IA pour analyser et configurer automatiquement."
        )
        desc.setStyleSheet("color: #a0a0c0; padding: 5px;")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Files section
        files_group = QGroupBox("📁 Fichiers Sources")
        files_layout = QVBoxLayout(files_group)
        
        # File buttons
        btn_layout = QHBoxLayout()
        self.add_files_btn = QPushButton("➕ Ajouter Fichiers")
        self.add_files_btn.clicked.connect(self._add_files)
        btn_layout.addWidget(self.add_files_btn)
        
        self.clear_files_btn = QPushButton("🗑️ Vider Liste")
        self.clear_files_btn.setObjectName("danger")
        self.clear_files_btn.clicked.connect(self._clear_files)
        btn_layout.addWidget(self.clear_files_btn)
        btn_layout.addStretch()
        files_layout.addLayout(btn_layout)
        
        # Files table
        self.files_table = QTableWidget()
        self.files_table.setColumnCount(5)
        self.files_table.setHorizontalHeaderLabels(["Fichier", "Responsable", "Branche", "Centre Coût", "Feuilles"])
        self.files_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.files_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.files_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.files_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.files_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.files_table.setColumnWidth(1, 150)
        self.files_table.setColumnWidth(2, 120)
        self.files_table.setColumnWidth(3, 120)
        self.files_table.setColumnWidth(4, 150)
        files_layout.addWidget(self.files_table)
        layout.addWidget(files_group)
        
        # Configuration section
        config_layout = QHBoxLayout()
        
        # Manual config
        config_group = QGroupBox("⚙️ Configuration Extraction")
        config_form = QFormLayout(config_group)
        
        # Global sheet name is less relevant now, but keeping as default/fallback
        self.sheet_name = QLineEdit("Feuille1")
        self.sheet_name.setPlaceholderText("Nom par défaut...")
        config_form.addRow("Feuille par déf.:", self.sheet_name)
        
        col_layout = QHBoxLayout()
        self.start_col = QLineEdit("E")
        self.start_col.setMaximumWidth(50)
        col_layout.addWidget(self.start_col)
        col_layout.addWidget(QLabel("à"))
        self.end_col = QLineEdit("P")
        self.end_col.setMaximumWidth(50)
        col_layout.addWidget(self.end_col)
        col_layout.addStretch()
        config_form.addRow("Colonnes:", col_layout)
        
        row_layout = QHBoxLayout()
        self.start_row = QSpinBox()
        self.start_row.setRange(1, 10000)
        self.start_row.setValue(1)
        row_layout.addWidget(self.start_row)
        row_layout.addWidget(QLabel("à"))
        self.end_row = QSpinBox()
        self.end_row.setRange(1, 10000)
        self.end_row.setValue(10)
        row_layout.addWidget(self.end_row)
        row_layout.addStretch()
        config_form.addRow("Lignes:", row_layout)
        
        self.group_by = QComboBox()
        self.group_by.addItems(["Branche", "Responsable", "Centre de Coût"])
        config_form.addRow("Grouper par:", self.group_by)
        
        config_layout.addWidget(config_group)
        
        # AI section
        ai_group = QGroupBox("🤖 Analyse IA (Optionnel)")
        ai_layout = QVBoxLayout(ai_group)
        
        self.use_ai = QCheckBox("Utiliser l'IA pour analyser")
        ai_layout.addWidget(self.use_ai)
        
        self.ai_key = QLineEdit()
        self.ai_key.setPlaceholderText("Clé API Z.ai...")
        self.ai_key.setEchoMode(QLineEdit.EchoMode.Password)
        ai_layout.addWidget(self.ai_key)
        
        self.ai_description = QTextEdit()
        self.ai_description.setPlaceholderText(
            "Décrivez ce que vous voulez consolider...\n"
            "Ex: Consolider les budgets par branche, colonnes E à P contiennent les montants mensuels"
        )
        self.ai_description.setMaximumHeight(80)
        ai_layout.addWidget(self.ai_description)
        
        self.ai_analyze_btn = QPushButton("🔍 Analyser avec IA")
        self.ai_analyze_btn.clicked.connect(self._analyze_with_ai)
        ai_layout.addWidget(self.ai_analyze_btn)
        
        config_layout.addWidget(ai_group)
        layout.addLayout(config_layout)
        
        # Progress
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #a78bfa;")
        layout.addWidget(self.status_label)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setObjectName("secondary")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        self.consolidate_btn = QPushButton("📊 Consolider")
        self.consolidate_btn.setObjectName("success")
        self.consolidate_btn.clicked.connect(self._start_consolidation)
        btn_layout.addWidget(self.consolidate_btn)
        
        layout.addLayout(btn_layout)
    
    def _add_files(self):
        """Add Excel files to the list"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "Sélectionner des fichiers Excel",
            "", "Fichiers Excel (*.xlsx *.xls)"
        )
        
        for filepath in files:
            self._add_file_row(filepath)
    
    def _add_file_row(self, filepath: str, responsible: str = "", branch: str = "", cost_center: str = "", sheets: list = None):
        """Add a file row to the table"""
        import os
        
        row = self.files_table.rowCount()
        self.files_table.insertRow(row)
        
        # Filename
        self.files_table.setItem(row, 0, QTableWidgetItem(os.path.basename(filepath)))
        self.files_table.item(row, 0).setData(Qt.ItemDataRole.UserRole, filepath)
        
        # Responsible
        resp_edit = QLineEdit(responsible or f"Site {row + 1}")
        self.files_table.setCellWidget(row, 1, resp_edit)
        
        # Branch
        branch_edit = QLineEdit(branch or f"Branche {row + 1}")
        self.files_table.setCellWidget(row, 2, branch_edit)
        
        # Cost center
        cost_edit = QLineEdit(cost_center)
        cost_edit.setPlaceholderText("Centre...")
        self.files_table.setCellWidget(row, 3, cost_edit)
        
        # Sheets
        try:
            wb = load_workbook(filepath, read_only=True, keep_links=False)
            all_sheets = wb.sheetnames
            wb.close()
        except Exception:
            all_sheets = ["Feuille1"]
            
        selected_sheets = sheets if sheets else [all_sheets[0]] if all_sheets else []
        
        sheet_btn = QPushButton(f"{len(selected_sheets)} feuille(s)")
        sheet_btn.clicked.connect(lambda checked, r=row, s=all_sheets: self._select_sheets(r, s))
        
        # Store sheets data in button property
        sheet_btn.setProperty("selected_sheets", selected_sheets)
        
        self.files_table.setCellWidget(row, 4, sheet_btn)
        
        self.files.append(filepath)

    def _select_sheets(self, row, all_sheets):
        """Open dialog to select sheets"""
        btn = self.files_table.cellWidget(row, 4)
        current_selection = btn.property("selected_sheets")
        
        dialog = SheetSelectionDialog(all_sheets, current_selection, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_selection = dialog.get_selected()
            btn.setProperty("selected_sheets", new_selection)
            btn.setText(f"{len(new_selection)} feuille(s)")
    
    def _clear_files(self):
        """Clear all files"""
        self.files_table.setRowCount(0)
        self.files.clear()
    
    def _get_file_sources(self):
        """Get file sources from table"""
        sources = []
        for row in range(self.files_table.rowCount()):
            filepath = self.files_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
            responsible = self.files_table.cellWidget(row, 1).text()
            branch = self.files_table.cellWidget(row, 2).text()
            cost_center = self.files_table.cellWidget(row, 3).text()
            
            sheet_btn = self.files_table.cellWidget(row, 4)
            sheets = sheet_btn.property("selected_sheets")
            
            sources.append(FileSource(
                filepath=filepath,
                responsible=responsible,
                branch=branch,
                cost_center=cost_center,
                sheets=sheets
            ))
        return sources
    
    def _analyze_with_ai(self):
        """Analyze files with AI"""
        if not self.files:
            QMessageBox.warning(self, "Erreur", "Ajoutez d'abord des fichiers.")
            return
        
        api_key = self.ai_key.text().strip()
        if not api_key:
            QMessageBox.warning(self, "Erreur", "Entrez votre clé API Z.ai.")
            return
        
        description = self.ai_description.toPlainText().strip() or "Consolider les fichiers Excel"
        
        self.progress_bar.setVisible(True)
        self.status_label.setText("🔄 Analyse IA en cours...")
        self.ai_analyze_btn.setEnabled(False)
        
        self.ai_worker = AIAnalysisWorker(self.files, description, api_key)
        self.ai_worker.finished.connect(self._on_ai_analysis_done)
        self.ai_worker.start()
    
    def _on_ai_analysis_done(self, result):
        """Handle AI analysis result"""
        self.progress_bar.setVisible(False)
        self.ai_analyze_btn.setEnabled(True)
        
        if isinstance(result, Exception):
            self.status_label.setText(f"❌ Erreur: {result}")
            return
        
        try:
            # Apply AI suggestions
            if "sheet_name" in result:
                self.sheet_name.setText(result["sheet_name"])
            if "start_column" in result:
                self.start_col.setText(result["start_column"])
            if "end_column" in result:
                self.end_col.setText(result["end_column"])
            if "start_row" in result:
                self.start_row.setValue(result["start_row"])
            if "end_row" in result:
                self.end_row.setValue(result["end_row"])
            
            # Update file info if provided
            if "files" in result:
                self._clear_files()
                for f in result["files"]:
                    self._add_file_row(
                        f.get("filepath", ""),
                        f.get("responsible", ""),
                        f.get("branch", ""),
                        f.get("cost_center", ""),
                        f.get("sheets", [])
                    )
            
            suggestions = result.get("suggestions", "Configuration appliquée!")
            self.status_label.setText(f"✅ IA: {suggestions[:100]}...")
            
        except Exception as e:
            self.status_label.setText(f"❌ Erreur: {e}")
    
    def _start_consolidation(self):
        """Start consolidation process"""
        sources = self._get_file_sources()
        
        if not sources:
            QMessageBox.warning(self, "Erreur", "Ajoutez au moins un fichier.")
            return
        
        # Get output path
        output_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le fichier consolidé",
            "Consolidation_BP_2026.xlsx", "Fichiers Excel (*.xlsx)"
        )
        
        if not output_path:
            return
        
        # Configure consolidator
        group_map = {0: "branch", 1: "responsible", 2: "cost_center"}
        
        config = ConsolidationConfig(
            sheet_name=self.sheet_name.text(),
            start_column=self.start_col.text().upper(),
            end_column=self.end_col.text().upper(),
            start_row=self.start_row.value(),
            end_row=self.end_row.value(),
            group_by=group_map.get(self.group_by.currentIndex(), "branch")
        )
        
        self.consolidator.set_config(config)
        self.consolidator.clear_sources()
        
        for source in sources:
            self.consolidator.add_source(source)
        
        # Start consolidation
        self.progress_bar.setVisible(True)
        self.consolidate_btn.setEnabled(False)
        self.status_label.setText("🔄 Consolidation en cours...")
        
        self.worker = ConsolidationWorker(self.consolidator, output_path)
        self.worker.finished.connect(self._on_consolidation_done)
        self.worker.start()
    
    def _on_consolidation_done(self, result: ConsolidationResult):
        """Handle consolidation result"""
        self.progress_bar.setVisible(False)
        self.consolidate_btn.setEnabled(True)
        
        if result.success:
            self.status_label.setText(f"✅ Consolidation réussie!")
            QMessageBox.information(
                self, "Succès",
                f"Consolidation terminée!\n\n"
                f"📁 Fichiers traités: {result.files_processed}\n"
                f"📋 Lignes extraites: {result.rows_extracted}\n"
                f"💾 Fichier: {result.output_path}"
            )
            self.accept()
        else:
            self.status_label.setText(f"❌ Erreur: {result.error}")
            QMessageBox.critical(self, "Erreur", f"Erreur: {result.error}")
