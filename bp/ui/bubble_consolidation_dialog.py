"""
Bubble Consolidation Dialog for PyQt6
Interactive bubble-based Excel consolidation
"""
import json
import os
from pathlib import Path
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QLabel,
    QFrame, QScrollArea, QFileDialog, QMessageBox, QSplitter,
    QTreeWidget, QTreeWidgetItem, QListWidget, QListWidgetItem,
    QProgressBar, QGroupBox
)
from PyQt6.QtCore import Qt, pyqtSignal, QMimeData
from PyQt6.QtGui import QDrag, QColor, QFont

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font as XLFont, PatternFill, Alignment


class BubbleWidget(QFrame):
    """Interactive bubble widget"""
    clicked = pyqtSignal(object)
    
    def __init__(self, data: dict, bubble_type: str, parent=None):
        super().__init__(parent)
        self.data = data
        self.bubble_type = bubble_type
        self._setup_ui()
        
    def _setup_ui(self):
        self.setFrameStyle(QFrame.Shape.StyledPanel)
        self.setCursor(Qt.CursorShape.OpenHandCursor)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        
        # Icon and text
        icons = {
            'responsable': '👤',
            'site': '📄',
            'sheet': '📋',
            'column': '📊'
        }
        
        colors = {
            'responsable': ('qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #3b82f6, stop:1 #1d4ed8)', 'white'),
            'site': ('qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #10b981, stop:1 #059669)', 'white'),
            'sheet': ('qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f59e0b, stop:1 #d97706)', 'white'),
            'column': ('qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #8b5cf6, stop:1 #7c3aed)', 'white')
        }
        
        bg, fg = colors.get(self.bubble_type, ('#e2e8f0', '#1e293b'))
        
        self.setStyleSheet(f"""
            QFrame {{
                background: {bg};
                border-radius: 16px;
                border: none;
            }}
            QFrame:hover {{
                border: 2px solid white;
            }}
            QLabel {{
                color: {fg};
                background: transparent;
            }}
        """)
        
        icon = QLabel(icons.get(self.bubble_type, '📦'))
        icon.setStyleSheet("font-size: 16px;")
        layout.addWidget(icon)
        
        text = QLabel(self.data.get('name', 'Item'))
        text.setStyleSheet("font-weight: 500; font-size: 13px;")
        layout.addWidget(text)
        
        # Count badge for responsable/site
        if self.bubble_type == 'responsable' and 'sites' in self.data:
            count = QLabel(f"{len(self.data['sites'])} sites")
            count.setStyleSheet("font-size: 11px; opacity: 0.8;")
            layout.addWidget(count)
        elif self.bubble_type == 'site' and 'sheets' in self.data:
            count = QLabel(f"{len(self.data['sheets'])} feuilles")
            count.setStyleSheet("font-size: 11px; opacity: 0.8;")
            layout.addWidget(count)
    
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit(self.data)
        super().mousePressEvent(event)
    
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.MouseButton.LeftButton:
            drag = QDrag(self)
            mime = QMimeData()
            mime.setText(json.dumps({
                'type': self.bubble_type,
                'data': self.data
            }))
            drag.setMimeData(mime)
            drag.exec(Qt.DropAction.CopyAction)


class DropZoneWidget(QFrame):
    """Drop zone for bubbles"""
    itemDropped = pyqtSignal(dict)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.dropped_items = []
        self._setup_ui()
        
    def _setup_ui(self):
        self.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border: 2px dashed #cbd5e1;
                border-radius: 12px;
                min-height: 300px;
            }
        """)
        
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(16, 16, 16, 16)
        self.layout.setSpacing(8)
        
        # Placeholder
        self.placeholder = QLabel("🎯 Glissez les bulles ici")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setStyleSheet("color: #94a3b8; font-size: 14px; padding: 60px;")
        self.layout.addWidget(self.placeholder)
        self.layout.addStretch()
    
    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            self.setStyleSheet("""
                QFrame {
                    background: #e0f2fe;
                    border: 2px dashed #0ea5e9;
                    border-radius: 12px;
                    min-height: 300px;
                }
            """)
            event.acceptProposedAction()
    
    def dragLeaveEvent(self, event):
        self.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border: 2px dashed #cbd5e1;
                border-radius: 12px;
                min-height: 300px;
            }
        """)
    
    def dropEvent(self, event):
        self.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border: 2px dashed #cbd5e1;
                border-radius: 12px;
                min-height: 300px;
            }
        """)
        
        try:
            data = json.loads(event.mimeData().text())
            self.add_dropped_item(data)
            self.itemDropped.emit(data)
        except:
            pass
    
    def add_dropped_item(self, data: dict):
        # Hide placeholder
        self.placeholder.hide()
        
        # Check if already exists
        for item in self.dropped_items:
            if item.get('data', {}).get('id') == data.get('data', {}).get('id'):
                return
        
        self.dropped_items.append(data)
        
        # Create visual item
        item_widget = QFrame()
        item_widget.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px;
            }
        """)
        item_layout = QHBoxLayout(item_widget)
        item_layout.setContentsMargins(12, 8, 12, 8)
        
        icons = {'responsable': '👤', 'site': '📄', 'sheet': '📋', 'column': '📊'}
        icon = QLabel(icons.get(data.get('type', ''), '📦'))
        item_layout.addWidget(icon)
        
        info = QVBoxLayout()
        name = QLabel(data.get('data', {}).get('name', 'Item'))
        name.setStyleSheet("font-weight: bold; color: #1e293b;")
        info.addWidget(name)
        
        type_label = QLabel(data.get('type', '').capitalize())
        type_label.setStyleSheet("font-size: 11px; color: #64748b;")
        info.addWidget(type_label)
        item_layout.addLayout(info)
        
        item_layout.addStretch()
        
        # Remove button
        remove_btn = QPushButton("✕")
        remove_btn.setFixedSize(24, 24)
        remove_btn.setStyleSheet("""
            QPushButton {
                background: #fef2f2;
                border: none;
                border-radius: 12px;
                color: #ef4444;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #fee2e2;
            }
        """)
        remove_btn.clicked.connect(lambda: self.remove_item(item_widget, data))
        item_layout.addWidget(remove_btn)
        
        # Insert before stretch
        self.layout.insertWidget(self.layout.count() - 1, item_widget)
    
    def remove_item(self, widget, data):
        widget.deleteLater()
        self.dropped_items = [i for i in self.dropped_items 
                             if i.get('data', {}).get('id') != data.get('data', {}).get('id')]
        
        if not self.dropped_items:
            self.placeholder.show()
    
    def clear(self):
        self.dropped_items = []
        for i in reversed(range(self.layout.count() - 1)):
            widget = self.layout.itemAt(i).widget()
            if widget and widget != self.placeholder:
                widget.deleteLater()
        self.placeholder.show()


class BubbleConsolidationDialog(QDialog):
    """Main dialog for bubble-based consolidation"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.bubble_data = {'responsables': []}
        self.files = {}  # filename -> file path
        self._setup_ui()
    
    def _setup_ui(self):
        self.setWindowTitle("🔮 Consolidation par Bulles")
        self.setMinimumSize(1200, 700)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QLabel("🔮 Consolidation par Bulles")
        header.setStyleSheet("font-size: 22px; font-weight: bold; color: #0284c7;")
        layout.addWidget(header)
        
        desc = QLabel("Visualisez et consolidez vos fichiers Excel de manière interactive avec des bulles")
        desc.setStyleSheet("color: #64748b; margin-bottom: 16px;")
        layout.addWidget(desc)
        
        # Main content splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel - Explorer
        left_panel = QFrame()
        left_panel.setStyleSheet("QFrame { background: white; border-radius: 12px; }")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(16, 16, 16, 16)
        
        # Folder selection
        folder_group = QGroupBox("📂 Charger un dossier")
        folder_layout = QVBoxLayout(folder_group)
        
        self.folder_btn = QPushButton("📁 Sélectionner un dossier de responsables")
        self.folder_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #0ea5e9, stop:1 #0284c7);
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 8px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #38bdf8, stop:1 #0ea5e9);
            }
        """)
        self.folder_btn.clicked.connect(self._select_folder)
        folder_layout.addWidget(self.folder_btn)
        
        left_layout.addWidget(folder_group)
        
        # Bubble tree
        tree_group = QGroupBox("🔮 Hiérarchie des Bulles")
        tree_layout = QVBoxLayout(tree_group)
        
        self.bubble_scroll = QScrollArea()
        self.bubble_scroll.setWidgetResizable(True)
        self.bubble_scroll.setStyleSheet("QScrollArea { border: none; background: #f8fafc; }")
        
        self.bubble_container = QWidget()
        self.bubble_layout = QVBoxLayout(self.bubble_container)
        self.bubble_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        self.empty_label = QLabel("📂 Chargez un dossier pour voir les bulles")
        self.empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.empty_label.setStyleSheet("color: #94a3b8; padding: 40px;")
        self.bubble_layout.addWidget(self.empty_label)
        
        self.bubble_scroll.setWidget(self.bubble_container)
        tree_layout.addWidget(self.bubble_scroll)
        
        left_layout.addWidget(tree_group)
        splitter.addWidget(left_panel)
        
        # Right panel - Drop zone
        right_panel = QFrame()
        right_panel.setStyleSheet("QFrame { background: white; border-radius: 12px; }")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(16, 16, 16, 16)
        
        design_header = QHBoxLayout()
        design_title = QLabel("📊 Zone de Conception")
        design_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        design_header.addWidget(design_title)
        design_header.addStretch()
        
        clear_btn = QPushButton("🗑️ Vider")
        clear_btn.setObjectName("secondary")
        clear_btn.clicked.connect(self._clear_design)
        design_header.addWidget(clear_btn)
        right_layout.addLayout(design_header)
        
        self.drop_zone = DropZoneWidget()
        self.drop_zone.itemDropped.connect(self._on_item_dropped)
        right_layout.addWidget(self.drop_zone)
        
        # Generate button
        self.generate_btn = QPushButton("📥 Générer le fichier Excel")
        self.generate_btn.setEnabled(False)
        self.generate_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #10b981, stop:1 #059669);
                color: white;
                border: none;
                padding: 14px 28px;
                border-radius: 8px;
                font-weight: 600;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #34d399, stop:1 #10b981);
            }
            QPushButton:disabled {
                background: #cbd5e1;
                color: #94a3b8;
            }
        """)
        self.generate_btn.clicked.connect(self._generate_excel)
        right_layout.addWidget(self.generate_btn)
        
        splitter.addWidget(right_panel)
        splitter.setSizes([500, 500])
        
        layout.addWidget(splitter)
        
        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("Fermer")
        close_btn.setObjectName("secondary")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
    
    def _select_folder(self):
        folder = QFileDialog.getExistingDirectory(
            self, "Sélectionner le dossier des responsables"
        )
        
        if folder:
            self._parse_folder(folder)
    
    def _parse_folder(self, folder_path: str):
        """Parse folder structure and create bubble data"""
        self.bubble_data = {'responsables': []}
        self.files = {}
        
        folder = Path(folder_path)
        
        # Look for subfolders (responsables) or Excel files
        for item in folder.iterdir():
            if item.is_dir():
                # This is a responsable folder
                resp_data = {
                    'id': f'resp_{len(self.bubble_data["responsables"])}',
                    'name': item.name.replace('_', ' '),
                    'sites': []
                }
                
                # Look for Excel files in this folder
                for excel_file in item.glob('*.xlsx'):
                    self._add_site_from_file(resp_data, excel_file)
                for excel_file in item.glob('*.xls'):
                    self._add_site_from_file(resp_data, excel_file)
                
                if resp_data['sites']:
                    self.bubble_data['responsables'].append(resp_data)
            
            elif item.suffix.lower() in ['.xlsx', '.xls']:
                # Excel file directly in folder
                if not self.bubble_data['responsables']:
                    self.bubble_data['responsables'].append({
                        'id': 'resp_0',
                        'name': folder.name.replace('_', ' '),
                        'sites': []
                    })
                self._add_site_from_file(self.bubble_data['responsables'][0], item)
        
        self._render_bubbles()
    
    def _add_site_from_file(self, resp_data: dict, file_path: Path):
        """Add a site from Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = []
            
            for idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                columns = []
                if ws.max_column:
                    for col_idx in range(1, min(ws.max_column + 1, 27)):
                        columns.append(get_column_letter(col_idx))
                
                sheets.append({
                    'id': f'sh_{len(resp_data["sites"])}_{idx}',
                    'name': sheet_name,
                    'columns': columns
                })
            
            wb.close()
            
            site_id = f'site_{len(resp_data["sites"])}'
            site_data = {
                'id': site_id,
                'name': file_path.stem.replace('_', ' '),
                'filename': file_path.name,
                'sheets': sheets
            }
            
            resp_data['sites'].append(site_data)
            self.files[file_path.name] = str(file_path)
            
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
    
    def _render_bubbles(self):
        """Render bubbles from data"""
        # Clear existing
        while self.bubble_layout.count():
            item = self.bubble_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if not self.bubble_data['responsables']:
            self.empty_label = QLabel("📂 Aucun fichier Excel trouvé")
            self.empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.empty_label.setStyleSheet("color: #94a3b8; padding: 40px;")
            self.bubble_layout.addWidget(self.empty_label)
            return
        
        for resp in self.bubble_data['responsables']:
            # Responsable bubble
            resp_bubble = BubbleWidget(resp, 'responsable')
            self.bubble_layout.addWidget(resp_bubble)
            
            # Sites container
            sites_container = QWidget()
            sites_container.setStyleSheet("margin-left: 24px;")
            sites_layout = QVBoxLayout(sites_container)
            sites_layout.setContentsMargins(24, 8, 0, 8)
            
            for site in resp.get('sites', []):
                site_bubble = BubbleWidget(site, 'site')
                sites_layout.addWidget(site_bubble)
                
                # Sheets container
                sheets_container = QWidget()
                sheets_layout = QVBoxLayout(sheets_container)
                sheets_layout.setContentsMargins(24, 4, 0, 4)
                
                for sheet in site.get('sheets', []):
                    sheet_bubble = BubbleWidget(sheet, 'sheet')
                    sheets_layout.addWidget(sheet_bubble)
                
                sites_layout.addWidget(sheets_container)
            
            self.bubble_layout.addWidget(sites_container)
        
        self.bubble_layout.addStretch()
    
    def _on_item_dropped(self, data: dict):
        self.generate_btn.setEnabled(True)
    
    def _clear_design(self):
        self.drop_zone.clear()
        self.generate_btn.setEnabled(False)
    
    def _generate_excel(self):
        """Generate consolidated Excel file"""
        if not self.drop_zone.dropped_items:
            return
        
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Sauvegarder le fichier consolidé",
            "Consolidation_Bulles.xlsx", "Fichiers Excel (*.xlsx)"
        )
        
        if not filepath:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Consolidation"
            
            current_row = 1
            
            for item in self.drop_zone.dropped_items:
                item_type = item.get('type', '')
                item_data = item.get('data', {})
                
                if item_type == 'responsable':
                    ws.cell(row=current_row, column=1, value=f"Responsable: {item_data.get('name', '')}")
                    ws.cell(row=current_row, column=1).font = XLFont(bold=True, size=14)
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                    current_row += 2
                    
                    # Add all sites for this responsable
                    for site in item_data.get('sites', []):
                        ws.cell(row=current_row, column=1, value=f"Site: {site.get('name', '')}")
                        ws.cell(row=current_row, column=1).font = XLFont(bold=True, size=12, color="059669")
                        current_row += 1
                        
                        # Load actual data
                        filename = site.get('filename', '')
                        if filename in self.files:
                            try:
                                src_wb = load_workbook(self.files[filename])
                                for sheet_name in src_wb.sheetnames[:2]:
                                    src_ws = src_wb[sheet_name]
                                    ws.cell(row=current_row, column=2, value=f"Feuille: {sheet_name}")
                                    current_row += 1
                                    
                                    for row_idx in range(1, min(11, src_ws.max_row + 1)):
                                        for col_idx in range(1, min(7, src_ws.max_column + 1)):
                                            val = src_ws.cell(row=row_idx, column=col_idx).value
                                            ws.cell(row=current_row, column=col_idx + 1, value=val)
                                        current_row += 1
                                    current_row += 1
                                src_wb.close()
                            except:
                                pass
                        current_row += 1
                
                elif item_type == 'site':
                    ws.cell(row=current_row, column=1, value=f"Site: {item_data.get('name', '')}")
                    ws.cell(row=current_row, column=1).font = XLFont(bold=True, size=12, color="059669")
                    current_row += 2
                
                elif item_type == 'sheet':
                    ws.cell(row=current_row, column=1, value=f"Feuille: {item_data.get('name', '')}")
                    ws.cell(row=current_row, column=1).font = XLFont(bold=True, color="D97706")
                    current_row += 2
            
            wb.save(filepath)
            wb.close()
            
            QMessageBox.information(
                self, "Succès",
                f"Fichier Excel généré avec succès!\n\n{filepath}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération: {e}")
