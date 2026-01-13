"""
Main window for the Excel Creator application
"""
import json
import os
from pathlib import Path
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QListWidget, QListWidgetItem, QPushButton, QLabel, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QSpinBox, QDoubleSpinBox,
    QLineEdit, QCheckBox, QComboBox, QColorDialog, QFileDialog,
    QMessageBox, QStatusBar, QToolBar, QFrame, QGroupBox, QFormLayout,
    QScrollArea, QGridLayout, QApplication, QStackedWidget
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QAction, QIcon, QColor, QFont

from .styles import DARK_THEME, HEADER_STYLE, CARD_STYLE
from core.models import (
    WorkbookConfig, SheetConfig, CellContent, CellStyle,
    MergeRange, ColumnConfig, RowConfig, Alignment, VerticalAlignment, BorderStyle
)
from core.excel_generator import ExcelGenerator
from ui.ai_dialog import AIPromptDialog
from ui.consolidation_dialog import ConsolidationDialog
from ui.bubble_consolidation_dialog import BubbleConsolidationDialog



class ColorButton(QPushButton):
    """Custom button for color selection"""
    colorChanged = pyqtSignal(str)

    def __init__(self, color: str = "FFFFFF", parent=None):
        super().__init__(parent)
        self._color = color
        self.setObjectName("colorButton")
        self.setFixedSize(50, 30)
        self.clicked.connect(self._pick_color)
        self._update_style()

    def _pick_color(self):
        color = QColorDialog.getColor(QColor(f"#{self._color}"), self, "Choisir une couleur")
        if color.isValid():
            self._color = color.name()[1:].upper()
            self._update_style()
            self.colorChanged.emit(self._color)

    def _update_style(self):
        self.setStyleSheet(f"background-color: #{self._color}; border: 2px solid #3d3d5c; border-radius: 6px;")

    def get_color(self) -> str:
        return self._color

    def set_color(self, color: str):
        self._color = color.upper()
        self._update_style()


class SheetListWidget(QWidget):
    """Widget for managing sheet list"""
    sheetSelected = pyqtSignal(int)
    sheetAdded = pyqtSignal()
    sheetRemoved = pyqtSignal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Title
        title = QLabel("📋 FEUILLES")
        title.setObjectName("section-title")
        layout.addWidget(title)

        # List
        self.list_widget = QListWidget()
        self.list_widget.currentRowChanged.connect(self.sheetSelected.emit)
        layout.addWidget(self.list_widget)

        # Buttons
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Ajouter")
        self.add_btn.clicked.connect(self.sheetAdded.emit)
        btn_layout.addWidget(self.add_btn)

        self.remove_btn = QPushButton("🗑️")
        self.remove_btn.setObjectName("danger")
        self.remove_btn.setFixedWidth(50)
        self.remove_btn.clicked.connect(self._remove_current)
        btn_layout.addWidget(self.remove_btn)

        layout.addLayout(btn_layout)

    def _remove_current(self):
        row = self.list_widget.currentRow()
        if row >= 0:
            self.sheetRemoved.emit(row)

    def add_sheet(self, name: str):
        item = QListWidgetItem(f"📄 {name}")
        self.list_widget.addItem(item)
        self.list_widget.setCurrentItem(item)

    def remove_sheet(self, index: int):
        self.list_widget.takeItem(index)

    def update_sheet_name(self, index: int, name: str):
        item = self.list_widget.item(index)
        if item:
            item.setText(f"📄 {name}")

    def clear(self):
        self.list_widget.clear()


class ColumnConfigWidget(QWidget):
    """Widget for configuring columns"""
    configChanged = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._columns = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Add column section
        add_layout = QHBoxLayout()
        
        self.col_count_spin = QSpinBox()
        self.col_count_spin.setRange(1, 100)
        self.col_count_spin.setValue(1)
        add_layout.addWidget(QLabel("Nombre:"))
        add_layout.addWidget(self.col_count_spin)
        add_layout.addStretch()

        self.add_col_btn = QPushButton("➕ Ajouter colonnes")
        self.add_col_btn.clicked.connect(self._add_columns)
        add_layout.addWidget(self.add_col_btn)

        layout.addLayout(add_layout)

        # Table for columns
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Colonne", "Largeur", "En-tête", "Actions"])
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(3, 80)
        layout.addWidget(self.table)

    def _add_columns(self):
        count = self.col_count_spin.value()
        start = len(self._columns) + 1
        
        for i in range(count):
            col_idx = start + i
            self._add_column_row(col_idx)
        
        self.configChanged.emit()

    def _add_column_row(self, col_idx: int):
        from openpyxl.utils import get_column_letter
        
        row = self.table.rowCount()
        self.table.insertRow(row)

        # Column letter
        col_letter = get_column_letter(col_idx)
        self.table.setItem(row, 0, QTableWidgetItem(col_letter))

        # Width spinbox
        width_spin = QDoubleSpinBox()
        width_spin.setRange(1, 100)
        width_spin.setValue(12.0)
        width_spin.valueChanged.connect(self.configChanged.emit)
        self.table.setCellWidget(row, 1, width_spin)

        # Header text
        header_edit = QLineEdit()
        header_edit.setPlaceholderText("Nom de la colonne...")
        header_edit.textChanged.connect(self.configChanged.emit)
        self.table.setCellWidget(row, 2, header_edit)

        # Delete button
        del_btn = QPushButton("🗑️")
        del_btn.setObjectName("danger")
        del_btn.clicked.connect(lambda: self._remove_column(row))
        self.table.setCellWidget(row, 3, del_btn)

        self._columns.append({
            'index': col_idx,
            'width_spin': width_spin,
            'header_edit': header_edit
        })

    def _remove_column(self, row: int):
        if row < len(self._columns):
            self.table.removeRow(row)
            self._columns.pop(row)
            self.configChanged.emit()

    def get_columns(self) -> list:
        columns = []
        for i, col_data in enumerate(self._columns):
            if i < self.table.rowCount():
                width_spin = self.table.cellWidget(i, 1)
                header_edit = self.table.cellWidget(i, 2)
                if width_spin and header_edit:
                    columns.append(ColumnConfig(
                        index=col_data['index'],
                        width=width_spin.value(),
                        header=header_edit.text()
                    ))
        return columns

    def set_columns(self, columns: list):
        self.table.setRowCount(0)
        self._columns.clear()
        for col in columns:
            self._add_column_row(col.index)
            row = self.table.rowCount() - 1
            self.table.cellWidget(row, 1).setValue(col.width)
            self.table.cellWidget(row, 2).setText(col.header)

    def clear(self):
        self.table.setRowCount(0)
        self._columns.clear()


class RowConfigWidget(QWidget):
    """Widget for configuring rows"""
    configChanged = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._rows = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Add row section
        add_layout = QHBoxLayout()
        
        self.row_count_spin = QSpinBox()
        self.row_count_spin.setRange(1, 1000)
        self.row_count_spin.setValue(1)
        add_layout.addWidget(QLabel("Nombre:"))
        add_layout.addWidget(self.row_count_spin)
        add_layout.addStretch()

        self.add_row_btn = QPushButton("➕ Ajouter lignes")
        self.add_row_btn.clicked.connect(self._add_rows)
        add_layout.addWidget(self.add_row_btn)

        layout.addLayout(add_layout)

        # Table for rows
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Ligne", "Hauteur", "Actions"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(2, 80)
        layout.addWidget(self.table)

    def _add_rows(self):
        count = self.row_count_spin.value()
        start = len(self._rows) + 1
        
        for i in range(count):
            row_idx = start + i
            self._add_row_entry(row_idx)
        
        self.configChanged.emit()

    def _add_row_entry(self, row_idx: int):
        row = self.table.rowCount()
        self.table.insertRow(row)

        # Row number
        self.table.setItem(row, 0, QTableWidgetItem(str(row_idx)))

        # Height spinbox
        height_spin = QDoubleSpinBox()
        height_spin.setRange(5, 500)
        height_spin.setValue(15.0)
        height_spin.valueChanged.connect(self.configChanged.emit)
        self.table.setCellWidget(row, 1, height_spin)

        # Delete button
        del_btn = QPushButton("🗑️")
        del_btn.setObjectName("danger")
        del_btn.clicked.connect(lambda: self._remove_row(row))
        self.table.setCellWidget(row, 2, del_btn)

        self._rows.append({
            'index': row_idx,
            'height_spin': height_spin
        })

    def _remove_row(self, row: int):
        if row < len(self._rows):
            self.table.removeRow(row)
            self._rows.pop(row)
            self.configChanged.emit()

    def get_rows(self) -> list:
        rows = []
        for i, row_data in enumerate(self._rows):
            if i < self.table.rowCount():
                height_spin = self.table.cellWidget(i, 1)
                if height_spin:
                    rows.append(RowConfig(
                        index=row_data['index'],
                        height=height_spin.value()
                    ))
        return rows

    def set_rows(self, rows: list):
        self.table.setRowCount(0)
        self._rows.clear()
        for row in rows:
            self._add_row_entry(row.index)
            table_row = self.table.rowCount() - 1
            self.table.cellWidget(table_row, 1).setValue(row.height)

    def clear(self):
        self.table.setRowCount(0)
        self._rows.clear()


class CellConfigWidget(QWidget):
    """Widget for configuring cell content and style"""
    configChanged = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._cells = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Add cell section
        add_group = QGroupBox("Ajouter une cellule")
        add_layout = QFormLayout(add_group)

        self.cell_row_spin = QSpinBox()
        self.cell_row_spin.setRange(1, 10000)
        self.cell_row_spin.setValue(1)
        add_layout.addRow("Ligne:", self.cell_row_spin)

        self.cell_col_spin = QSpinBox()
        self.cell_col_spin.setRange(1, 1000)
        self.cell_col_spin.setValue(1)
        add_layout.addRow("Colonne:", self.cell_col_spin)

        self.cell_value = QLineEdit()
        self.cell_value.setPlaceholderText("Contenu de la cellule...")
        add_layout.addRow("Valeur:", self.cell_value)

        # Style options
        style_layout = QHBoxLayout()
        
        self.bold_check = QCheckBox("Gras")
        style_layout.addWidget(self.bold_check)
        
        self.italic_check = QCheckBox("Italique")
        style_layout.addWidget(self.italic_check)
        
        self.underline_check = QCheckBox("Souligné")
        style_layout.addWidget(self.underline_check)
        
        add_layout.addRow("Style:", style_layout)

        # Colors
        color_layout = QHBoxLayout()
        
        color_layout.addWidget(QLabel("Fond:"))
        self.bg_color_btn = ColorButton("FFFFFF")
        color_layout.addWidget(self.bg_color_btn)
        
        color_layout.addWidget(QLabel("Texte:"))
        self.font_color_btn = ColorButton("000000")
        color_layout.addWidget(self.font_color_btn)
        
        color_layout.addStretch()
        add_layout.addRow("Couleurs:", color_layout)

        # Font size
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(6, 72)
        self.font_size_spin.setValue(11)
        add_layout.addRow("Taille police:", self.font_size_spin)

        # Alignment
        self.align_combo = QComboBox()
        self.align_combo.addItems(["Gauche", "Centre", "Droite"])
        add_layout.addRow("Alignement:", self.align_combo)

        # Border
        self.border_combo = QComboBox()
        self.border_combo.addItems(["Aucune", "Fine", "Moyenne", "Épaisse", "Double"])
        add_layout.addRow("Bordure:", self.border_combo)

        # Add button
        self.add_cell_btn = QPushButton("➕ Ajouter cellule")
        self.add_cell_btn.clicked.connect(self._add_cell)
        add_layout.addRow("", self.add_cell_btn)

        layout.addWidget(add_group)

        # Cells table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Cellule", "Valeur", "Style", "Couleurs", "Actions"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

    def _add_cell(self):
        row = self.cell_row_spin.value()
        col = self.cell_col_spin.value()
        value = self.cell_value.text()

        # Create style
        align_map = {0: Alignment.LEFT, 1: Alignment.CENTER, 2: Alignment.RIGHT}
        border_map = {
            0: BorderStyle.NONE, 1: BorderStyle.THIN,
            2: BorderStyle.MEDIUM, 3: BorderStyle.THICK, 4: BorderStyle.DOUBLE
        }

        style = CellStyle(
            bold=self.bold_check.isChecked(),
            italic=self.italic_check.isChecked(),
            underline=self.underline_check.isChecked(),
            bg_color=self.bg_color_btn.get_color() if self.bg_color_btn.get_color() != "FFFFFF" else None,
            font_color=self.font_color_btn.get_color(),
            font_size=self.font_size_spin.value(),
            alignment=align_map.get(self.align_combo.currentIndex(), Alignment.LEFT),
            border_style=border_map.get(self.border_combo.currentIndex(), BorderStyle.NONE)
        )

        cell = CellContent(row=row, col=col, value=value, style=style)
        self._cells.append(cell)
        self._add_cell_to_table(cell)
        
        # Clear inputs
        self.cell_value.clear()
        self.configChanged.emit()

    def _add_cell_to_table(self, cell: CellContent):
        from openpyxl.utils import get_column_letter
        
        row = self.table.rowCount()
        self.table.insertRow(row)

        # Cell reference
        cell_ref = f"{get_column_letter(cell.col)}{cell.row}"
        self.table.setItem(row, 0, QTableWidgetItem(cell_ref))

        # Value
        self.table.setItem(row, 1, QTableWidgetItem(cell.value))

        # Style description
        style_parts = []
        if cell.style.bold:
            style_parts.append("B")
        if cell.style.italic:
            style_parts.append("I")
        if cell.style.underline:
            style_parts.append("U")
        style_parts.append(f"{cell.style.font_size}pt")
        self.table.setItem(row, 2, QTableWidgetItem(" ".join(style_parts)))

        # Colors
        color_text = f"#{cell.style.font_color}"
        if cell.style.bg_color:
            color_text += f" sur #{cell.style.bg_color}"
        self.table.setItem(row, 3, QTableWidgetItem(color_text))

        # Delete button
        del_btn = QPushButton("🗑️")
        del_btn.setObjectName("danger")
        del_btn.clicked.connect(lambda: self._remove_cell(row))
        self.table.setCellWidget(row, 4, del_btn)

    def _remove_cell(self, row: int):
        if row < len(self._cells):
            self.table.removeRow(row)
            self._cells.pop(row)
            self.configChanged.emit()

    def get_cells(self) -> list:
        return self._cells.copy()

    def set_cells(self, cells: list):
        self.table.setRowCount(0)
        self._cells.clear()
        for cell in cells:
            self._cells.append(cell)
            self._add_cell_to_table(cell)

    def clear(self):
        self.table.setRowCount(0)
        self._cells.clear()


class MergeConfigWidget(QWidget):
    """Widget for configuring cell merges"""
    configChanged = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._merges = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Add merge section
        add_group = QGroupBox("Ajouter une fusion")
        add_layout = QGridLayout(add_group)

        add_layout.addWidget(QLabel("Début:"), 0, 0)
        add_layout.addWidget(QLabel("Ligne:"), 0, 1)
        self.start_row = QSpinBox()
        self.start_row.setRange(1, 10000)
        add_layout.addWidget(self.start_row, 0, 2)
        add_layout.addWidget(QLabel("Colonne:"), 0, 3)
        self.start_col = QSpinBox()
        self.start_col.setRange(1, 1000)
        add_layout.addWidget(self.start_col, 0, 4)

        add_layout.addWidget(QLabel("Fin:"), 1, 0)
        add_layout.addWidget(QLabel("Ligne:"), 1, 1)
        self.end_row = QSpinBox()
        self.end_row.setRange(1, 10000)
        add_layout.addWidget(self.end_row, 1, 2)
        add_layout.addWidget(QLabel("Colonne:"), 1, 3)
        self.end_col = QSpinBox()
        self.end_col.setRange(1, 1000)
        add_layout.addWidget(self.end_col, 1, 4)

        self.add_merge_btn = QPushButton("➕ Ajouter fusion")
        self.add_merge_btn.clicked.connect(self._add_merge)
        add_layout.addWidget(self.add_merge_btn, 2, 0, 1, 5)

        layout.addWidget(add_group)

        # Merges table
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Plage", "Cellules fusionnées", "Actions"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

    def _add_merge(self):
        merge = MergeRange(
            start_row=self.start_row.value(),
            start_col=self.start_col.value(),
            end_row=self.end_row.value(),
            end_col=self.end_col.value()
        )

        if merge.end_row < merge.start_row or merge.end_col < merge.start_col:
            QMessageBox.warning(self, "Erreur", "La cellule de fin doit être après la cellule de début.")
            return

        self._merges.append(merge)
        self._add_merge_to_table(merge)
        self.configChanged.emit()

    def _add_merge_to_table(self, merge: MergeRange):
        from openpyxl.utils import get_column_letter

        row = self.table.rowCount()
        self.table.insertRow(row)

        # Range
        range_str = merge.to_excel_range()
        self.table.setItem(row, 0, QTableWidgetItem(range_str))

        # Cell count
        cell_count = (merge.end_row - merge.start_row + 1) * (merge.end_col - merge.start_col + 1)
        self.table.setItem(row, 1, QTableWidgetItem(f"{cell_count} cellules"))

        # Delete button
        del_btn = QPushButton("🗑️")
        del_btn.setObjectName("danger")
        del_btn.clicked.connect(lambda: self._remove_merge(row))
        self.table.setCellWidget(row, 2, del_btn)

    def _remove_merge(self, row: int):
        if row < len(self._merges):
            self.table.removeRow(row)
            self._merges.pop(row)
            self.configChanged.emit()

    def get_merges(self) -> list:
        return self._merges.copy()

    def set_merges(self, merges: list):
        self.table.setRowCount(0)
        self._merges.clear()
        for merge in merges:
            self._merges.append(merge)
            self._add_merge_to_table(merge)

    def clear(self):
        self.table.setRowCount(0)
        self._merges.clear()


class PreviewWidget(QWidget):
    """Widget for previewing the Excel structure"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        # Title
        title = QLabel("👁️ APERÇU")
        title.setObjectName("section-title")
        layout.addWidget(title)

        # Preview table
        self.table = QTableWidget()
        self.table.setRowCount(10)
        self.table.setColumnCount(5)
        layout.addWidget(self.table)

    def update_preview(self, sheet_config: SheetConfig):
        """Update preview based on sheet configuration"""
        # Determine size
        max_row = 10
        max_col = 5

        for cell in sheet_config.cells:
            max_row = max(max_row, cell.row + 1)
            max_col = max(max_col, cell.col + 1)

        for merge in sheet_config.merges:
            max_row = max(max_row, merge.end_row + 1)
            max_col = max(max_col, merge.end_col + 1)

        self.table.setRowCount(max_row)
        self.table.setColumnCount(max_col)

        # Set headers
        from openpyxl.utils import get_column_letter
        headers = [get_column_letter(i + 1) for i in range(max_col)]
        self.table.setHorizontalHeaderLabels(headers)

        # Clear previous content
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                self.table.setItem(r, c, QTableWidgetItem(""))

        # Apply column widths
        for col in sheet_config.columns:
            if col.index <= max_col:
                self.table.setColumnWidth(col.index - 1, int(col.width * 8))
                if col.header:
                    item = QTableWidgetItem(col.header)
                    item.setBackground(QColor("#4472C4"))
                    item.setForeground(QColor("white"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    self.table.setItem(0, col.index - 1, item)

        # Apply cells
        for cell in sheet_config.cells:
            item = QTableWidgetItem(cell.value)
            
            if cell.style.bg_color:
                item.setBackground(QColor(f"#{cell.style.bg_color}"))
            if cell.style.font_color:
                item.setForeground(QColor(f"#{cell.style.font_color}"))
            
            font = QFont()
            font.setBold(cell.style.bold)
            font.setItalic(cell.style.italic)
            font.setUnderline(cell.style.underline)
            font.setPointSize(cell.style.font_size)
            item.setFont(font)

            self.table.setItem(cell.row - 1, cell.col - 1, item)

        # Visual indication for merges
        for merge in sheet_config.merges:
            for r in range(merge.start_row, merge.end_row + 1):
                for c in range(merge.start_col, merge.end_col + 1):
                    item = self.table.item(r - 1, c - 1)
                    if not item:
                        item = QTableWidgetItem("")
                        self.table.setItem(r - 1, c - 1, item)
                    item.setBackground(QColor("#3d3d5c"))
                    if r == merge.start_row and c == merge.start_col:
                        item.setText("🔗")



class MainWindow(QMainWindow):
    """Main application window with sidebar navigation"""

    def __init__(self):
        super().__init__()
        self.config = WorkbookConfig()
        self.current_sheet_index = 0
        self.project_file = None
        self._load_app_settings()
        self._setup_ui()
        self._connect_signals()
        self._refresh_sheets()

    def _load_app_settings(self):
        """Load app settings from file"""
        import os
        self.settings_file = os.path.join(os.path.dirname(__file__), '..', 'app_settings.json')
        self.app_name = "Excel Tools"
        self.app_subtitle = "BP 2026"
        self.logo_url = "https://pereire.co/wp-content/uploads/2023/06/Logo.svg"
        
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.app_name = settings.get('app_name', self.app_name)
                    self.app_subtitle = settings.get('app_subtitle', self.app_subtitle)
                    self.logo_url = settings.get('logo_url', self.logo_url)
        except:
            pass

    def _setup_ui(self):
        self.setWindowTitle("📊 Excel Tools - BP 2026")
        self.setMinimumSize(1300, 850)
        self.setStyleSheet(DARK_THEME)

        # Central widget
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # === SIDEBAR ===
        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(220)
        sidebar.setStyleSheet("""
            QWidget#sidebar {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0284c7, stop:1 #0369a1);
            }
        """)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(16, 20, 16, 20)
        sidebar_layout.setSpacing(8)

        header_layout = QHBoxLayout()
        logo = QLabel("📊")
        logo.setStyleSheet("font-size: 28px;")
        header_layout.addWidget(logo)
        
        title_layout = QVBoxLayout()
        self.sidebar_title = QLabel(self.app_name)
        self.sidebar_title.setStyleSheet("font-size: 18px; font-weight: bold; color: white;")
        title_layout.addWidget(self.sidebar_title)
        self.sidebar_subtitle = QLabel(self.app_subtitle)
        self.sidebar_subtitle.setStyleSheet("font-size: 11px; color: rgba(255,255,255,0.7);")
        title_layout.addWidget(self.sidebar_subtitle)
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        sidebar_layout.addLayout(header_layout)
        
        sidebar_layout.addSpacing(30)

        # Navigation buttons
        self.nav_consolidation = QPushButton("📁  Consolidation")
        self.nav_consolidation.setObjectName("nav-item")
        self.nav_consolidation.setCheckable(True)
        self.nav_consolidation.setChecked(True)
        self.nav_consolidation.setStyleSheet("""
            QPushButton#nav-item {
                background: rgba(255,255,255,0.2);
                color: white;
                border: none;
                border-radius: 10px;
                padding: 14px 16px;
                text-align: left;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton#nav-item:hover {
                background: rgba(255,255,255,0.15);
            }
            QPushButton#nav-item:checked {
                background: rgba(255,255,255,0.25);
            }
        """)
        self.nav_consolidation.clicked.connect(lambda: self._switch_page(0))
        sidebar_layout.addWidget(self.nav_consolidation)

        self.nav_creator = QPushButton("📄  Excel Creator")
        self.nav_creator.setObjectName("nav-item")
        self.nav_creator.setCheckable(True)
        self.nav_creator.setStyleSheet("""
            QPushButton#nav-item {
                background: transparent;
                color: rgba(255,255,255,0.85);
                border: none;
                border-radius: 10px;
                padding: 14px 16px;
                text-align: left;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton#nav-item:hover {
                background: rgba(255,255,255,0.1);
                color: white;
            }
            QPushButton#nav-item:checked {
                background: rgba(255,255,255,0.2);
                color: white;
            }
        """)
        self.nav_creator.clicked.connect(lambda: self._switch_page(1))
        sidebar_layout.addWidget(self.nav_creator)

        # Bubbles nav
        self.nav_bubbles = QPushButton("🔮  Bulles")
        self.nav_bubbles.setObjectName("nav-item")
        self.nav_bubbles.setCheckable(True)
        self.nav_bubbles.setStyleSheet("""
            QPushButton#nav-item {
                background: transparent;
                color: rgba(255,255,255,0.85);
                border: none;
                border-radius: 10px;
                padding: 14px 16px;
                text-align: left;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton#nav-item:hover {
                background: rgba(255,255,255,0.1);
                color: white;
            }
            QPushButton#nav-item:checked {
                background: rgba(255,255,255,0.2);
                color: white;
            }
        """)
        self.nav_bubbles.clicked.connect(lambda: self._switch_page(2))
        sidebar_layout.addWidget(self.nav_bubbles)

        # ConsBulle nav
        self.nav_consbulle = QPushButton("🧩  Cons. par Bulle")
        self.nav_consbulle.setObjectName("nav-item")
        self.nav_consbulle.setCheckable(True)
        self.nav_consbulle.setStyleSheet("""
            QPushButton#nav-item {
                background: transparent;
                color: rgba(255,255,255,0.85);
                border: none;
                border-radius: 10px;
                padding: 14px 16px;
                text-align: left;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton#nav-item:hover {
                background: rgba(255,255,255,0.1);
                color: white;
            }
            QPushButton#nav-item:checked {
                background: rgba(255,255,255,0.2);
                color: white;
            }
        """)
        self.nav_consbulle.clicked.connect(lambda: self._switch_page(3))
        sidebar_layout.addWidget(self.nav_consbulle)

        # Settings nav
        self.nav_settings = QPushButton("⚙️  Paramètres")
        self.nav_settings.setObjectName("nav-item")
        self.nav_settings.setCheckable(True)
        self.nav_settings.setStyleSheet("""
            QPushButton#nav-item {
                background: transparent;
                color: rgba(255,255,255,0.85);
                border: none;
                border-radius: 10px;
                padding: 14px 16px;
                text-align: left;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton#nav-item:hover {
                background: rgba(255,255,255,0.1);
                color: white;
            }
            QPushButton#nav-item:checked {
                background: rgba(255,255,255,0.2);
                color: white;
            }
        """)
        self.nav_settings.clicked.connect(lambda: self._switch_page(4))
        sidebar_layout.addWidget(self.nav_settings)

        sidebar_layout.addStretch()

        # Version
        version = QLabel("v1.0")
        version.setStyleSheet("color: rgba(255,255,255,0.5); font-size: 11px;")
        version.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sidebar_layout.addWidget(version)

        main_layout.addWidget(sidebar)

        # === MAIN CONTENT AREA ===
        content_area = QWidget()
        content_area.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f0f9ff, stop:0.5 #e0f2fe, stop:1 #f0f9ff);")
        content_layout = QVBoxLayout(content_area)
        content_layout.setContentsMargins(0, 0, 0, 0)

        # Stacked widget for pages
        self.pages = QStackedWidget()
        
        # Page 1: Consolidation
        self.consolidation_page = self._create_consolidation_page()
        self.pages.addWidget(self.consolidation_page)
        
        # Page 2: Excel Creator
        self.creator_page = self._create_creator_page()
        self.pages.addWidget(self.creator_page)
        
        # Page 3: Bubbles
        self.bubbles_page = self._create_bubbles_page()
        self.pages.addWidget(self.bubbles_page)
        
        # Page 4: ConsBulle
        self.consbulle_page = self._create_consbulle_page()
        self.pages.addWidget(self.consbulle_page)
        
        # Page 5: Settings
        self.settings_page = self._create_settings_page()
        self.pages.addWidget(self.settings_page)

        content_layout.addWidget(self.pages)
        main_layout.addWidget(content_area)

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("✓ Prêt")
        
        # Update window title
        self.setWindowTitle(f"📊 {self.app_name} - {self.app_subtitle}")

    def _switch_page(self, index: int):
        self.pages.setCurrentIndex(index)
        self.nav_consolidation.setChecked(index == 0)
        self.nav_creator.setChecked(index == 1)
        self.nav_bubbles.setChecked(index == 2)
        self.nav_consbulle.setChecked(index == 3)
        self.nav_settings.setChecked(index == 4)

    def _create_consolidation_page(self):
        """Create the consolidation page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)

        # Header
        header_layout = QVBoxLayout()
        title = QLabel("📁 Consolidation de Fichiers Excel")
        title.setStyleSheet("font-size: 26px; font-weight: bold; color: #1e293b;")
        header_layout.addWidget(title)
        subtitle = QLabel("Combinez plusieurs fichiers Excel en un seul fichier de synthèse")
        subtitle.setStyleSheet("font-size: 14px; color: #64748b;")
        header_layout.addWidget(subtitle)
        layout.addLayout(header_layout)

        # Open Consolidation Dialog Button
        btn_layout = QHBoxLayout()
        self.open_consolidation_btn = QPushButton("📂 Ouvrir l'outil de consolidation")
        self.open_consolidation_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #0ea5e9, stop:1 #0284c7);
                color: white;
                border: none;
                padding: 16px 32px;
                border-radius: 12px;
                font-size: 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #38bdf8, stop:1 #0ea5e9);
            }
        """)
        self.open_consolidation_btn.clicked.connect(self._open_consolidation_dialog)
        btn_layout.addWidget(self.open_consolidation_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Info card
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
                padding: 24px;
            }
        """)
        info_layout = QVBoxLayout(info_card)
        
        info_title = QLabel("💡 Comment utiliser la consolidation ?")
        info_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #0284c7; margin-bottom: 12px;")
        info_layout.addWidget(info_title)
        
        steps = [
            "1. Cliquez sur le bouton ci-dessus pour ouvrir l'outil de consolidation",
            "2. Ajoutez les fichiers Excel à consolider",
            "3. Sélectionnez les feuilles à inclure pour chaque fichier",
            "4. Configurez les colonnes et lignes à extraire",
            "5. Lancez la consolidation et téléchargez le résultat"
        ]
        for step in steps:
            step_label = QLabel(step)
            step_label.setStyleSheet("font-size: 14px; color: #64748b; padding: 4px 0;")
            info_layout.addWidget(step_label)
        
        layout.addWidget(info_card)
        layout.addStretch()

        return page

    def _create_bubbles_page(self):
        """Create the bubbles consolidation page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)

        # Header
        header_layout = QVBoxLayout()
        title = QLabel("🔮 Consolidation par Bulles")
        title.setStyleSheet("font-size: 26px; font-weight: bold; color: #1e293b;")
        header_layout.addWidget(title)
        subtitle = QLabel("Visualisez et consolidez vos fichiers Excel de manière interactive avec des bulles")
        subtitle.setStyleSheet("font-size: 14px; color: #64748b;")
        header_layout.addWidget(subtitle)
        layout.addLayout(header_layout)

        # Open Bubble Dialog Button
        btn_layout = QHBoxLayout()
        self.open_bubbles_btn = QPushButton("🔮 Ouvrir l'outil de consolidation par bulles")
        self.open_bubbles_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #8b5cf6, stop:1 #7c3aed);
                color: white;
                border: none;
                padding: 16px 32px;
                border-radius: 12px;
                font-size: 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #a78bfa, stop:1 #8b5cf6);
            }
        """)
        self.open_bubbles_btn.clicked.connect(self._open_bubble_dialog)
        btn_layout.addWidget(self.open_bubbles_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Info card
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
                padding: 24px;
            }
        """)
        info_layout = QVBoxLayout(info_card)
        
        info_title = QLabel("💡 Comment fonctionne la consolidation par bulles ?")
        info_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #7c3aed; margin-bottom: 12px;")
        info_layout.addWidget(info_title)
        
        steps = [
            "1. Sélectionnez un dossier contenant les sous-dossiers des responsables",
            "2. Chaque responsable contient des fichiers Excel (un par site)",
            "3. Les fichiers sont visualisés sous forme de bulles interactives",
            "4. Glissez-déposez les bulles vers la zone de conception",
            "5. Générez le fichier Excel consolidé"
        ]
        for step in steps:
            step_label = QLabel(step)
            step_label.setStyleSheet("font-size: 14px; color: #64748b; padding: 4px 0;")
            info_layout.addWidget(step_label)
        
        # Structure example
        structure_label = QLabel("\n📁 Structure attendue:")
        structure_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #1e293b; margin-top: 12px;")
        info_layout.addWidget(structure_label)
        
        structure_example = QLabel("""
    Responsables/
    ├── Jean_Dupont/
    │   ├── Site_Paris.xlsx
    │   ├── Site_Lyon.xlsx
    │   └── Site_Marseille.xlsx
    └── Marie_Martin/
        ├── Site_Nice.xlsx
        └── Site_Lille.xlsx
        """)
        structure_example.setStyleSheet("font-family: monospace; font-size: 12px; color: #64748b; background: #f8fafc; padding: 12px; border-radius: 8px;")
        info_layout.addWidget(structure_example)
        
        layout.addWidget(info_card)
        layout.addStretch()

        return page

    def _open_bubble_dialog(self):
        """Open bubble consolidation dialog"""
        dialog = BubbleConsolidationDialog(self)
        dialog.exec()

    def _create_consbulle_page(self):
        """Create the structured consolidation by bubble page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)

        # Header
        header_layout = QVBoxLayout()
        title = QLabel("🧩 Consolidation par Bulle (Structurée)")
        title.setStyleSheet("font-size: 26px; font-weight: bold; color: #1e293b;")
        header_layout.addWidget(title)
        subtitle = QLabel("Assistant étape par étape pour créer une consolidation structurée")
        subtitle.setStyleSheet("font-size: 14px; color: #64748b;")
        header_layout.addWidget(subtitle)
        layout.addLayout(header_layout)

        # Open Dialog Button
        btn_layout = QHBoxLayout()
        self.open_consbulle_btn = QPushButton("🧩 Ouvrir l'assistant de consolidation structurée")
        self.open_consbulle_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f59e0b, stop:1 #d97706);
                color: white;
                border: none;
                padding: 16px 32px;
                border-radius: 12px;
                font-size: 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #fbbf24, stop:1 #f59e0b);
            }
        """)
        self.open_consbulle_btn.clicked.connect(self._open_consbulle_dialog)
        btn_layout.addWidget(self.open_consbulle_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Info card
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
                padding: 24px;
            }
        """)
        info_layout = QVBoxLayout(info_card)
        
        info_title = QLabel("💡 Comment fonctionne l'assistant ?")
        info_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #d97706; margin-bottom: 12px;")
        info_layout.addWidget(info_title)
        
        steps = [
            "👤 Étape 1 : Créez les responsables (ex: Jean Dupont, Marie Martin)",
            "📄 Étape 2 : Ajoutez des sites et chargez les fichiers Excel",
            "⚙️ Étape 3 : Configurez l'extraction (colonnes E-P, lignes 1-10)",
            "📊 Étape 4 : Aperçu et génération du fichier consolidé"
        ]
        for step in steps:
            step_label = QLabel(step)
            step_label.setStyleSheet("font-size: 14px; color: #64748b; padding: 6px 0;")
            info_layout.addWidget(step_label)
        
        layout.addWidget(info_card)
        layout.addStretch()

        return page

    def _open_consbulle_dialog(self):
        """Open structured consolidation dialog"""
        # For now, use the same bubble dialog - can be extended later
        dialog = BubbleConsolidationDialog(self)
        dialog.exec()

    def _create_settings_page(self):
        """Create the settings page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)

        # Header
        header_layout = QVBoxLayout()
        title = QLabel("⚙️ Paramètres")
        title.setStyleSheet("font-size: 26px; font-weight: bold; color: #1e293b;")
        header_layout.addWidget(title)
        subtitle = QLabel("Personnalisez l'apparence de l'application")
        subtitle.setStyleSheet("font-size: 14px; color: #64748b;")
        header_layout.addWidget(subtitle)
        layout.addLayout(header_layout)

        # Settings card
        settings_card = QFrame()
        settings_card.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
                padding: 24px;
            }
        """)
        settings_layout = QVBoxLayout(settings_card)
        
        card_title = QLabel("🏢 Branding")
        card_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #0284c7; margin-bottom: 16px;")
        settings_layout.addWidget(card_title)

        # App name
        name_group = QGroupBox("Nom de l'application")
        name_group.setStyleSheet("QGroupBox { font-weight: bold; color: #1e293b; }")
        name_layout = QVBoxLayout(name_group)
        self.settings_app_name = QLineEdit(self.app_name)
        self.settings_app_name.setPlaceholderText("Nom de l'application...")
        name_layout.addWidget(self.settings_app_name)
        settings_layout.addWidget(name_group)

        # App subtitle
        subtitle_group = QGroupBox("Sous-titre")
        subtitle_group.setStyleSheet("QGroupBox { font-weight: bold; color: #1e293b; }")
        subtitle_layout = QVBoxLayout(subtitle_group)
        self.settings_app_subtitle = QLineEdit(self.app_subtitle)
        self.settings_app_subtitle.setPlaceholderText("Sous-titre...")
        subtitle_layout.addWidget(self.settings_app_subtitle)
        settings_layout.addWidget(subtitle_group)

        # Logo URL
        logo_group = QGroupBox("URL du logo")
        logo_group.setStyleSheet("QGroupBox { font-weight: bold; color: #1e293b; }")
        logo_layout = QVBoxLayout(logo_group)
        self.settings_logo_url = QLineEdit(self.logo_url)
        self.settings_logo_url.setPlaceholderText("https://...")
        logo_layout.addWidget(self.settings_logo_url)
        logo_hint = QLabel("Formats supportés: SVG, PNG, JPG")
        logo_hint.setStyleSheet("font-size: 12px; color: #64748b;")
        logo_layout.addWidget(logo_hint)
        settings_layout.addWidget(logo_group)

        # Save button
        save_btn = QPushButton("💾 Sauvegarder les paramètres")
        save_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #0ea5e9, stop:1 #0284c7);
                color: white;
                border: none;
                padding: 14px 28px;
                border-radius: 10px;
                font-size: 14px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #38bdf8, stop:1 #0ea5e9);
            }
        """)
        save_btn.clicked.connect(self._save_app_settings)
        settings_layout.addWidget(save_btn)

        layout.addWidget(settings_card)
        layout.addStretch()

        return page

    def _save_app_settings(self):
        """Save app settings to file"""
        self.app_name = self.settings_app_name.text()
        self.app_subtitle = self.settings_app_subtitle.text()
        self.logo_url = self.settings_logo_url.text()
        
        settings = {
            'app_name': self.app_name,
            'app_subtitle': self.app_subtitle,
            'logo_url': self.logo_url
        }
        
        try:
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)
            
            # Update UI
            self.sidebar_title.setText(self.app_name)
            self.sidebar_subtitle.setText(self.app_subtitle)
            self.setWindowTitle(f"📊 {self.app_name} - {self.app_subtitle}")
            
            self.status_bar.showMessage("✓ Paramètres sauvegardés!")
            QMessageBox.information(self, "Succès", "Les paramètres ont été sauvegardés!")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde: {e}")

    def _create_creator_page(self):
        """Create the Excel Creator page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Toolbar
        self._create_toolbar_widget(layout)

        # Content area
        content = QWidget()
        content_layout = QHBoxLayout(content)
        content_layout.setContentsMargins(15, 15, 15, 15)
        content_layout.setSpacing(15)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel - Sheet list
        self.sheet_list = SheetListWidget()
        self.sheet_list.setMaximumWidth(250)
        splitter.addWidget(self.sheet_list)

        # Center panel - Configuration
        center_panel = QFrame()
        center_panel.setObjectName("panel")
        center_panel.setStyleSheet("QFrame#panel { background: white; border: 1px solid #e2e8f0; border-radius: 16px; }")
        center_layout = QVBoxLayout(center_panel)
        center_layout.setContentsMargins(15, 15, 15, 15)

        # Sheet name
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Nom de la feuille:"))
        self.sheet_name_edit = QLineEdit()
        self.sheet_name_edit.setPlaceholderText("Nom de la feuille...")
        name_layout.addWidget(self.sheet_name_edit)
        center_layout.addLayout(name_layout)

        # Tabs for configuration
        self.config_tabs = QTabWidget()
        
        self.column_config = ColumnConfigWidget()
        self.config_tabs.addTab(self.column_config, "📊 Colonnes")

        self.row_config = RowConfigWidget()
        self.config_tabs.addTab(self.row_config, "📋 Lignes")

        self.cell_config = CellConfigWidget()
        self.config_tabs.addTab(self.cell_config, "✏️ Cellules")

        self.merge_config = MergeConfigWidget()
        self.config_tabs.addTab(self.merge_config, "🔗 Fusions")

        center_layout.addWidget(self.config_tabs)
        splitter.addWidget(center_panel)

        # Right panel - Preview
        self.preview = PreviewWidget()
        self.preview.setMaximumWidth(350)
        splitter.addWidget(self.preview)

        splitter.setSizes([200, 600, 300])
        content_layout.addWidget(splitter)
        layout.addWidget(content)

        return page

    def _create_toolbar_widget(self, parent_layout):
        """Create toolbar for Excel Creator page"""
        from PyQt6.QtWidgets import QStackedWidget
        
        toolbar = QFrame()
        toolbar.setStyleSheet("""
            QFrame {
                background: white;
                border-bottom: 1px solid #e2e8f0;
                padding: 12px 20px;
            }
        """)
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(20, 12, 20, 12)

        # Page title
        title = QLabel("📄 Excel Creator")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1e293b;")
        toolbar_layout.addWidget(title)
        toolbar_layout.addStretch()

        # Buttons
        new_btn = QPushButton("📄 Nouveau")
        new_btn.setObjectName("secondary")
        new_btn.clicked.connect(self._new_project)
        toolbar_layout.addWidget(new_btn)

        open_btn = QPushButton("📂 Ouvrir")
        open_btn.setObjectName("secondary")
        open_btn.clicked.connect(self._open_project)
        toolbar_layout.addWidget(open_btn)

        save_btn = QPushButton("💾 Sauvegarder")
        save_btn.setObjectName("secondary")
        save_btn.clicked.connect(self._save_project)
        toolbar_layout.addWidget(save_btn)

        ai_btn = QPushButton("🤖 IA")
        ai_btn.setObjectName("ai-button")
        ai_btn.clicked.connect(self._open_ai_dialog)
        toolbar_layout.addWidget(ai_btn)

        generate_btn = QPushButton("📊 Générer Excel")
        generate_btn.clicked.connect(self._generate_excel)
        toolbar_layout.addWidget(generate_btn)

        parent_layout.addWidget(toolbar)

    def _create_toolbar(self):
        # Kept for compatibility but not used with new layout
        pass


    def _connect_signals(self):
        self.sheet_list.sheetSelected.connect(self._on_sheet_selected)
        self.sheet_list.sheetAdded.connect(self._add_sheet)
        self.sheet_list.sheetRemoved.connect(self._remove_sheet)

        self.sheet_name_edit.textChanged.connect(self._on_sheet_name_changed)
        
        self.column_config.configChanged.connect(self._on_config_changed)
        self.row_config.configChanged.connect(self._on_config_changed)
        self.cell_config.configChanged.connect(self._on_config_changed)
        self.merge_config.configChanged.connect(self._on_config_changed)

    def _refresh_sheets(self):
        self.sheet_list.clear()
        for sheet in self.config.sheets:
            self.sheet_list.add_sheet(sheet.name)
        
        if self.config.sheets:
            self._load_sheet_config(0)

    def _on_sheet_selected(self, index: int):
        if index >= 0 and index < len(self.config.sheets):
            self._save_current_sheet()
            self.current_sheet_index = index
            self._load_sheet_config(index)

    def _save_current_sheet(self):
        if self.current_sheet_index < len(self.config.sheets):
            sheet = self.config.sheets[self.current_sheet_index]
            sheet.name = self.sheet_name_edit.text() or f"Feuille{self.current_sheet_index + 1}"
            sheet.columns = self.column_config.get_columns()
            sheet.rows = self.row_config.get_rows()
            sheet.cells = self.cell_config.get_cells()
            sheet.merges = self.merge_config.get_merges()

    def _load_sheet_config(self, index: int):
        if index < len(self.config.sheets):
            sheet = self.config.sheets[index]
            self.sheet_name_edit.setText(sheet.name)
            self.column_config.set_columns(sheet.columns)
            self.row_config.set_rows(sheet.rows)
            self.cell_config.set_cells(sheet.cells)
            self.merge_config.set_merges(sheet.merges)
            self._update_preview()

    def _add_sheet(self):
        name = f"Feuille{len(self.config.sheets) + 1}"
        sheet = SheetConfig(name=name)
        self.config.sheets.append(sheet)
        self.sheet_list.add_sheet(name)
        self.status_bar.showMessage(f"✓ Feuille '{name}' ajoutée")

    def _remove_sheet(self, index: int):
        if len(self.config.sheets) <= 1:
            QMessageBox.warning(self, "Erreur", "Vous devez garder au moins une feuille.")
            return

        name = self.config.sheets[index].name
        self.config.sheets.pop(index)
        self.sheet_list.remove_sheet(index)
        
        if self.current_sheet_index >= len(self.config.sheets):
            self.current_sheet_index = len(self.config.sheets) - 1
        
        self._load_sheet_config(self.current_sheet_index)
        self.status_bar.showMessage(f"✓ Feuille '{name}' supprimée")

    def _on_sheet_name_changed(self, name: str):
        if self.current_sheet_index < len(self.config.sheets):
            self.sheet_list.update_sheet_name(self.current_sheet_index, name)

    def _on_config_changed(self):
        self._save_current_sheet()
        self._update_preview()

    def _update_preview(self):
        if self.current_sheet_index < len(self.config.sheets):
            sheet = self.config.sheets[self.current_sheet_index]
            self.preview.update_preview(sheet)

    def _new_project(self):
        reply = QMessageBox.question(
            self, "Nouveau projet",
            "Créer un nouveau projet ? Les modifications non sauvegardées seront perdues.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.config = WorkbookConfig()
            self.project_file = None
            self.current_sheet_index = 0
            self._refresh_sheets()
            self.status_bar.showMessage("✓ Nouveau projet créé")

    def _open_project(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Ouvrir un projet",
            "", "Fichiers JSON (*.json)"
        )
        
        if filepath:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.config = WorkbookConfig.from_dict(data)
                self.project_file = filepath
                self.current_sheet_index = 0
                self._refresh_sheets()
                self.status_bar.showMessage(f"✓ Projet chargé: {filepath}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Erreur lors de l'ouverture: {e}")

    def _save_project(self):
        self._save_current_sheet()
        
        if not self.project_file:
            filepath, _ = QFileDialog.getSaveFileName(
                self, "Sauvegarder le projet",
                "bp_2026_config.json", "Fichiers JSON (*.json)"
            )
            if not filepath:
                return
            self.project_file = filepath
        
        try:
            with open(self.project_file, 'w', encoding='utf-8') as f:
                json.dump(self.config.to_dict(), f, indent=2, ensure_ascii=False)
            self.status_bar.showMessage(f"✓ Projet sauvegardé: {self.project_file}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde: {e}")

    def _generate_excel(self):
        self._save_current_sheet()
        
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Générer le fichier Excel",
            "BP_2026.xlsx", "Fichiers Excel (*.xlsx)"
        )
        
        if filepath:
            try:
                generator = ExcelGenerator()
                if generator.generate(self.config, filepath):
                    self.status_bar.showMessage(f"✓ Fichier Excel généré: {filepath}")
                    QMessageBox.information(
                        self, "Succès",
                        f"Le fichier Excel a été généré avec succès!\n\n{filepath}"
                    )
                else:
                    QMessageBox.critical(self, "Erreur", "Erreur lors de la génération du fichier.")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération: {e}")

    def _open_ai_dialog(self):
        """Open AI prompt dialog"""
        dialog = AIPromptDialog(self)
        dialog.configGenerated.connect(self._apply_ai_config)
        dialog.exec()

    def _apply_ai_config(self, config: WorkbookConfig):
        """Apply configuration generated by AI"""
        self.config = config
        self.current_sheet_index = 0
        self._refresh_sheets()
        self.status_bar.showMessage("✓ Configuration IA appliquée avec succès!")

    def _open_consolidation_dialog(self):
        """Open consolidation dialog"""
        dialog = ConsolidationDialog(self)
        dialog.exec()
