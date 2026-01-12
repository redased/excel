"""
Excel File Consolidation Module
Combines multiple Excel files by sheet, columns, and rows
"""
import os
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



@dataclass
class FileSource:
    """Source file configuration"""
    filepath: str
    responsible: str  # Responsable/Site name
    branch: str = ""  # Branche
    cost_center: str = ""  # Centre de coût
    sheets: List[str] = field(default_factory=list)  # Sheets to consolidate


@dataclass
class ConsolidationConfig:
    """Configuration for consolidation"""
    sheet_name: str = "Feuille1"  # Default fallback
    start_column: str = "E"  # Starting column letter
    end_column: str = "P"    # Ending column letter
    start_row: int = 1
    end_row: int = 10
    include_headers: bool = True
    group_by: str = "branch"  # branch, responsible, cost_center


@dataclass
class ConsolidationResult:
    """Result of consolidation"""
    success: bool
    output_path: Optional[str] = None
    files_processed: int = 0
    rows_extracted: int = 0
    error: Optional[str] = None


class ExcelConsolidator:
    """Consolidates multiple Excel files into one summary file"""

    def __init__(self):
        self.sources: List[FileSource] = []
        self.config = ConsolidationConfig()

    def add_source(self, source: FileSource):
        """Add a source file"""
        self.sources.append(source)

    def clear_sources(self):
        """Clear all sources"""
        self.sources.clear()

    def set_config(self, config: ConsolidationConfig):
        """Set consolidation configuration"""
        self.config = config

    def consolidate(self, output_path: str) -> ConsolidationResult:
        """
        Consolidate all source files into one output file
        
        Groups data by the configured group_by field
        """
        if not self.sources:
            return ConsolidationResult(
                success=False,
                error="Aucun fichier source ajouté"
            )

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Consolidation"

            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Calculate column range
            start_col_idx = column_index_from_string(self.config.start_column)
            end_col_idx = column_index_from_string(self.config.end_column)
            num_cols = end_col_idx - start_col_idx + 1

            # Write main headers
            current_row = 1
            headers = ["Source", "Feuille", "Responsable", "Branche", "Centre de Coût"]
            
            # Add column headers from range
            for i in range(num_cols):
                headers.append(f"Col {get_column_letter(start_col_idx + i)}")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border

            current_row += 1
            total_rows = 0

            # Group sources
            groups: Dict[str, List[FileSource]] = {}
            for source in self.sources:
                key = getattr(source, self.config.group_by, source.responsible)
                if key not in groups:
                    groups[key] = []
                groups[key].append(source)

            # Process each source file
            for group_name, group_sources in groups.items():
                # Add group header
                group_cell = ws.cell(row=current_row, column=1, value=f"📁 {group_name}")
                group_cell.font = Font(bold=True, color="ED7D31")
                ws.merge_cells(start_row=current_row, start_column=1, 
                             end_row=current_row, end_column=len(headers))
                current_row += 1

                for source in group_sources:
                    try:
                        if not os.path.exists(source.filepath):
                            continue

                        source_wb = load_workbook(source.filepath, data_only=True)
                        
                        # Determine sheets to process
                        sheets_to_process = source.sheets if source.sheets else [self.config.sheet_name]
                        
                        for sheet_name in sheets_to_process:
                            if sheet_name not in source_wb.sheetnames:
                                continue

                            source_ws = source_wb[sheet_name]

                            # Extract data from specified range
                            for row in range(self.config.start_row, self.config.end_row + 1):
                                # Check if row has data
                                has_data = False
                                row_data = []
                                
                                for col in range(start_col_idx, end_col_idx + 1):
                                    cell_value = source_ws.cell(row=row, column=col).value
                                    row_data.append(cell_value)
                                    if cell_value is not None:
                                        has_data = True

                                if has_data:
                                    # Write source info
                                    ws.cell(row=current_row, column=1, value=os.path.basename(source.filepath))
                                    ws.cell(row=current_row, column=2, value=sheet_name)
                                    ws.cell(row=current_row, column=3, value=source.responsible)
                                    ws.cell(row=current_row, column=4, value=source.branch)
                                    ws.cell(row=current_row, column=5, value=source.cost_center)

                                    # Write extracted data
                                    for col_offset, value in enumerate(row_data):
                                        cell = ws.cell(row=current_row, column=6 + col_offset, value=value)
                                        cell.border = border

                                    current_row += 1
                                    total_rows += 1

                        source_wb.close()

                    except Exception as e:
                        # Log error but continue with other files
                        print(f"Error processing {source.filepath}: {e}")
                        continue

                # Add empty row between groups
                current_row += 1

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Save output
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

            return ConsolidationResult(
                success=True,
                output_path=output_path,
                files_processed=len(self.sources),
                rows_extracted=total_rows
            )

        except Exception as e:
            return ConsolidationResult(
                success=False,
                error=str(e)
            )


def consolidate_files(
    files: List[Dict],
    sheet_name: str,
    start_column: str,
    end_column: str,
    start_row: int,
    end_row: int,
    output_path: str,
    group_by: str = "branch"
) -> ConsolidationResult:
    """
    Convenience function to consolidate files
    
    Args:
        files: List of dicts with keys: filepath, responsible, branch, cost_center, sheets (list)
        sheet_name: Default sheet to extract from if not specified in file
        ...
    """
    consolidator = ExcelConsolidator()
    
    config = ConsolidationConfig(
        sheet_name=sheet_name,
        start_column=start_column,
        end_column=end_column,
        start_row=start_row,
        end_row=end_row,
        group_by=group_by
    )
    consolidator.set_config(config)
    
    for file_info in files:
        source = FileSource(
            filepath=file_info.get("filepath", ""),
            responsible=file_info.get("responsible", ""),
            branch=file_info.get("branch", ""),
            cost_center=file_info.get("cost_center", ""),
            sheets=file_info.get("sheets", [])
        )
        consolidator.add_source(source)
    
    return consolidator.consolidate(output_path)
