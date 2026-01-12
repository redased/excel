"""
Data models for Excel configuration
"""
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any
from enum import Enum


class Alignment(Enum):
    LEFT = "left"
    CENTER = "center"
    RIGHT = "right"


class VerticalAlignment(Enum):
    TOP = "top"
    CENTER = "center"
    BOTTOM = "bottom"


class BorderStyle(Enum):
    NONE = None
    THIN = "thin"
    MEDIUM = "medium"
    THICK = "thick"
    DOUBLE = "double"


@dataclass
class CellStyle:
    """Style configuration for a cell"""
    font_name: str = "Calibri"
    font_size: int = 11
    font_color: str = "000000"  # Hex color without #
    bg_color: Optional[str] = None  # Hex color without #
    bold: bool = False
    italic: bool = False
    underline: bool = False
    alignment: Alignment = Alignment.LEFT
    vertical_alignment: VerticalAlignment = VerticalAlignment.CENTER
    border_style: BorderStyle = BorderStyle.NONE
    border_color: str = "000000"
    wrap_text: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "font_name": self.font_name,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "bg_color": self.bg_color,
            "bold": self.bold,
            "italic": self.italic,
            "underline": self.underline,
            "alignment": self.alignment.value,
            "vertical_alignment": self.vertical_alignment.value,
            "border_style": self.border_style.value,
            "border_color": self.border_color,
            "wrap_text": self.wrap_text
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "CellStyle":
        return cls(
            font_name=data.get("font_name", "Calibri"),
            font_size=data.get("font_size", 11),
            font_color=data.get("font_color", "000000"),
            bg_color=data.get("bg_color"),
            bold=data.get("bold", False),
            italic=data.get("italic", False),
            underline=data.get("underline", False),
            alignment=Alignment(data.get("alignment", "left")),
            vertical_alignment=VerticalAlignment(data.get("vertical_alignment", "center")),
            border_style=BorderStyle(data.get("border_style")),
            border_color=data.get("border_color", "000000"),
            wrap_text=data.get("wrap_text", False)
        )


@dataclass
class CellContent:
    """Content and style for a specific cell"""
    row: int  # 1-indexed
    col: int  # 1-indexed
    value: str = ""
    style: CellStyle = field(default_factory=CellStyle)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "row": self.row,
            "col": self.col,
            "value": self.value,
            "style": self.style.to_dict()
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "CellContent":
        return cls(
            row=data["row"],
            col=data["col"],
            value=data.get("value", ""),
            style=CellStyle.from_dict(data.get("style", {}))
        )


@dataclass
class MergeRange:
    """Definition of a cell merge range"""
    start_row: int  # 1-indexed
    start_col: int  # 1-indexed
    end_row: int    # 1-indexed
    end_col: int    # 1-indexed

    def to_dict(self) -> Dict[str, int]:
        return {
            "start_row": self.start_row,
            "start_col": self.start_col,
            "end_row": self.end_row,
            "end_col": self.end_col
        }

    @classmethod
    def from_dict(cls, data: Dict[str, int]) -> "MergeRange":
        return cls(
            start_row=data["start_row"],
            start_col=data["start_col"],
            end_row=data["end_row"],
            end_col=data["end_col"]
        )

    def to_excel_range(self) -> str:
        """Convert to Excel range notation (e.g., A1:B2)"""
        from openpyxl.utils import get_column_letter
        start = f"{get_column_letter(self.start_col)}{self.start_row}"
        end = f"{get_column_letter(self.end_col)}{self.end_row}"
        return f"{start}:{end}"


@dataclass
class ColumnConfig:
    """Configuration for a column"""
    index: int  # 1-indexed
    width: float = 12.0
    header: str = ""
    header_style: CellStyle = field(default_factory=lambda: CellStyle(bold=True, bg_color="4472C4", font_color="FFFFFF"))

    def to_dict(self) -> Dict[str, Any]:
        return {
            "index": self.index,
            "width": self.width,
            "header": self.header,
            "header_style": self.header_style.to_dict()
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ColumnConfig":
        return cls(
            index=data["index"],
            width=data.get("width", 12.0),
            header=data.get("header", ""),
            header_style=CellStyle.from_dict(data.get("header_style", {}))
        )


@dataclass
class RowConfig:
    """Configuration for a row"""
    index: int  # 1-indexed
    height: float = 15.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "index": self.index,
            "height": self.height
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "RowConfig":
        return cls(
            index=data["index"],
            height=data.get("height", 15.0)
        )


@dataclass
class SheetConfig:
    """Complete configuration for a worksheet"""
    name: str = "Feuille1"
    columns: List[ColumnConfig] = field(default_factory=list)
    rows: List[RowConfig] = field(default_factory=list)
    cells: List[CellContent] = field(default_factory=list)
    merges: List[MergeRange] = field(default_factory=list)
    default_column_width: float = 12.0
    default_row_height: float = 15.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "columns": [c.to_dict() for c in self.columns],
            "rows": [r.to_dict() for r in self.rows],
            "cells": [c.to_dict() for c in self.cells],
            "merges": [m.to_dict() for m in self.merges],
            "default_column_width": self.default_column_width,
            "default_row_height": self.default_row_height
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "SheetConfig":
        return cls(
            name=data.get("name", "Feuille1"),
            columns=[ColumnConfig.from_dict(c) for c in data.get("columns", [])],
            rows=[RowConfig.from_dict(r) for r in data.get("rows", [])],
            cells=[CellContent.from_dict(c) for c in data.get("cells", [])],
            merges=[MergeRange.from_dict(m) for m in data.get("merges", [])],
            default_column_width=data.get("default_column_width", 12.0),
            default_row_height=data.get("default_row_height", 15.0)
        )


@dataclass
class WorkbookConfig:
    """Complete workbook configuration"""
    sheets: List[SheetConfig] = field(default_factory=lambda: [SheetConfig()])
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheets": [s.to_dict() for s in self.sheets]
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "WorkbookConfig":
        return cls(
            sheets=[SheetConfig.from_dict(s) for s in data.get("sheets", [])]
        )
