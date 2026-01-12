"""
Excel file generator using openpyxl
"""
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    WorkbookConfig, SheetConfig, CellContent, CellStyle,
    MergeRange, ColumnConfig, RowConfig, BorderStyle,
    Alignment as CellAlignment, VerticalAlignment
)


class ExcelGenerator:
    """Generates Excel files from configuration"""

    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.config: Optional[WorkbookConfig] = None

    def create_workbook(self, config: WorkbookConfig) -> Workbook:
        """Create a workbook from configuration"""
        self.config = config
        self.workbook = Workbook()
        
        # Remove default sheet
        if self.workbook.active:
            self.workbook.remove(self.workbook.active)

        # Add configured sheets
        for sheet_config in config.sheets:
            self._add_sheet(sheet_config)

        return self.workbook

    def _add_sheet(self, sheet_config: SheetConfig):
        """Add a worksheet with configuration"""
        ws = self.workbook.create_sheet(title=sheet_config.name)

        # Set default dimensions
        ws.sheet_format.defaultColWidth = sheet_config.default_column_width
        ws.sheet_format.defaultRowHeight = sheet_config.default_row_height

        # Apply column configurations
        for col_config in sheet_config.columns:
            self._apply_column_config(ws, col_config)

        # Apply row configurations
        for row_config in sheet_config.rows:
            self._apply_row_config(ws, row_config)

        # Apply cell contents and styles
        for cell_content in sheet_config.cells:
            self._apply_cell_content(ws, cell_content)

        # Apply merges
        for merge in sheet_config.merges:
            self._apply_merge(ws, merge)

    def _apply_column_config(self, ws, col_config: ColumnConfig):
        """Apply column configuration"""
        col_letter = get_column_letter(col_config.index)
        ws.column_dimensions[col_letter].width = col_config.width

        # Set header if specified
        if col_config.header:
            cell = ws.cell(row=1, column=col_config.index)
            cell.value = col_config.header
            self._apply_cell_style(cell, col_config.header_style)

    def _apply_row_config(self, ws, row_config: RowConfig):
        """Apply row configuration"""
        ws.row_dimensions[row_config.index].height = row_config.height

    def _apply_cell_content(self, ws, cell_content: CellContent):
        """Apply cell content and style"""
        cell = ws.cell(row=cell_content.row, column=cell_content.col)
        cell.value = cell_content.value
        self._apply_cell_style(cell, cell_content.style)

    def _apply_cell_style(self, cell, style: CellStyle):
        """Apply style to a cell"""
        # Font
        cell.font = Font(
            name=style.font_name,
            size=style.font_size,
            color=style.font_color,
            bold=style.bold,
            italic=style.italic,
            underline='single' if style.underline else None
        )

        # Background fill
        if style.bg_color:
            cell.fill = PatternFill(
                start_color=style.bg_color,
                end_color=style.bg_color,
                fill_type='solid'
            )

        # Alignment
        cell.alignment = Alignment(
            horizontal=style.alignment.value,
            vertical=style.vertical_alignment.value,
            wrap_text=style.wrap_text
        )

        # Border
        if style.border_style != BorderStyle.NONE:
            side = Side(
                style=style.border_style.value,
                color=style.border_color
            )
            cell.border = Border(
                left=side,
                right=side,
                top=side,
                bottom=side
            )

    def _apply_merge(self, ws, merge: MergeRange):
        """Apply cell merge"""
        ws.merge_cells(
            start_row=merge.start_row,
            start_column=merge.start_col,
            end_row=merge.end_row,
            end_column=merge.end_col
        )

    def save(self, filepath: str) -> bool:
        """Save the workbook to file"""
        if not self.workbook:
            raise ValueError("No workbook created. Call create_workbook first.")
        
        try:
            # Ensure directory exists
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(filepath)
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False

    def generate(self, config: WorkbookConfig, filepath: str) -> bool:
        """Create and save workbook in one step"""
        self.create_workbook(config)
        return self.save(filepath)
