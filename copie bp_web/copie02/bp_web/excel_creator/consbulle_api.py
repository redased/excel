"""
API Views for Consolidation par Bulle with SQLite persistence
"""
import json
import os
from pathlib import Path

from django.http import JsonResponse, FileResponse
from django.views import View
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.numbers import FORMAT_NUMBER_00

from .models import ConsolidationConfig, Responsable, Site, SiteSheetConfig


# ============================================
# CONFIG CRUD APIs
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class ConsolidationConfigListAPIView(View):
    """List and create consolidation configs"""
    
    def get(self, request):
        """List all saved configurations"""
        configs = ConsolidationConfig.objects.all()
        return JsonResponse({
            'success': True,
            'configs': [
                {
                    'id': c.id,
                    'name': c.name,
                    'output_filename': c.output_filename,
                    'updated_at': c.updated_at.isoformat(),
                    'responsables_count': c.responsables.count()
                }
                for c in configs
            ]
        })
    
    def post(self, request):
        """Create a new configuration"""
        try:
            data = json.loads(request.body)
            config = ConsolidationConfig.objects.create(
                name=data.get('name', 'Nouvelle Configuration'),
                output_filename=data.get('output_filename', 'Consolidation'),
                selected_sheets=data.get('selected_sheets', [])
            )
            return JsonResponse({
                'success': True,
                'config': config.to_dict()
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)


@method_decorator(csrf_exempt, name='dispatch')
class ConsolidationConfigDetailAPIView(View):
    """Get, update, delete a specific config"""
    
    def get(self, request, config_id):
        """Get full configuration with all relationships"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            return JsonResponse({
                'success': True,
                'config': config.to_dict()
            })
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)
    
    def put(self, request, config_id):
        """Update configuration"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            data = json.loads(request.body)
            
            if 'name' in data:
                config.name = data['name']
            if 'output_filename' in data:
                config.output_filename = data['output_filename']
            if 'selected_sheets' in data:
                config.selected_sheets = data['selected_sheets']
            
            config.save()
            return JsonResponse({'success': True, 'config': config.to_dict()})
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    def delete(self, request, config_id):
        """Delete configuration"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            config.delete()
            return JsonResponse({'success': True})
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)


# ============================================
# RESPONSABLE APIs
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class ResponsableAPIView(View):
    """Create, update, delete responsables"""
    
    def post(self, request, config_id):
        """Add a responsable to a config"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            data = json.loads(request.body)
            
            resp = Responsable.objects.create(
                config=config,
                name=data.get('name', 'Nouveau Responsable'),
                order=config.responsables.count()
            )
            return JsonResponse({'success': True, 'responsable': resp.to_dict()})
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)
    
    def put(self, request, config_id, resp_id):
        """Update a responsable"""
        try:
            resp = Responsable.objects.get(id=resp_id, config_id=config_id)
            data = json.loads(request.body)
            
            if 'name' in data:
                resp.name = data['name']
            if 'order' in data:
                resp.order = data['order']
            
            resp.save()
            return JsonResponse({'success': True, 'responsable': resp.to_dict()})
        except Responsable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Responsable not found'}, status=404)
    
    def delete(self, request, config_id, resp_id):
        """Delete a responsable"""
        try:
            resp = Responsable.objects.get(id=resp_id, config_id=config_id)
            resp.delete()
            return JsonResponse({'success': True})
        except Responsable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Responsable not found'}, status=404)


# ============================================
# SITE APIs
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class SiteAPIView(View):
    """Create, update, delete sites"""
    
    def post(self, request, resp_id):
        """Add a site to a responsable"""
        try:
            resp = Responsable.objects.get(id=resp_id)
            data = json.loads(request.body)
            
            site = Site.objects.create(
                responsable=resp,
                name=data.get('name', 'Nouveau Site'),
                original_filename=data.get('original_filename', ''),
                detected_sheets=data.get('detected_sheets', []),
                order=resp.sites.count()
            )
            return JsonResponse({'success': True, 'site': site.to_dict()})
        except Responsable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Responsable not found'}, status=404)
    
    def put(self, request, resp_id, site_id):
        """Update a site"""
        try:
            site = Site.objects.get(id=site_id, responsable_id=resp_id)
            data = json.loads(request.body)
            
            if 'name' in data:
                site.name = data['name']
            if 'detected_sheets' in data:
                site.detected_sheets = data['detected_sheets']
            if 'order' in data:
                site.order = data['order']
            
            site.save()
            return JsonResponse({'success': True, 'site': site.to_dict()})
        except Site.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Site not found'}, status=404)
    
    def delete(self, request, resp_id, site_id):
        """Delete a site"""
        try:
            site = Site.objects.get(id=site_id, responsable_id=resp_id)
            site.delete()
            return JsonResponse({'success': True})
        except Site.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Site not found'}, status=404)


# ============================================
# PARSE EXCEL SHEETS API
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class ParseExcelSheetsAPIView(View):
    """Parse Excel files and return sheet names"""
    
    def post(self, request):
        """Parse uploaded files and return sheet names"""
        try:
            files = request.FILES.getlist('files')
            all_sheets = set()
            file_sheets = {}
            
            for f in files:
                try:
                    wb = load_workbook(f, read_only=True)
                    sheets = wb.sheetnames
                    file_sheets[f.name] = sheets
                    all_sheets.update(sheets)
                    wb.close()
                except Exception as e:
                    file_sheets[f.name] = {'error': str(e)}
            
            return JsonResponse({
                'success': True,
                'all_sheets': sorted(list(all_sheets)),
                'file_sheets': file_sheets
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# GENERATE EXCEL WITH FORMAT PRESERVATION
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class GenerateConsBulleV2APIView(View):
    """Generate consolidated Excel with format preservation"""
    
    def post(self, request):
        try:
            config_str = request.POST.get('config', '{}')
            config_data = json.loads(config_str)
            output_filename = config_data.get('output_filename', 'Consolidation')
            files = request.FILES.getlist('files')
            
            # Parse file mappings
            file_mappings = request.POST.getlist('file_mapping')
            file_map = {}
            for mapping_str in file_mappings:
                mapping = json.loads(mapping_str)
                file_map[mapping['filename']] = mapping
            
            # Match files to mappings
            uploaded_files = {}
            for f in files:
                uploaded_files[f.name] = f
            
            # Create output workbook
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Get selected sheets
            selected_sheets = config_data.get('selected_sheets', [])
            
            # Process each responsable
            for resp in config_data.get('responsables', []):
                for site in resp.get('sites', []):
                    filename = site.get('filename')
                    
                    if filename and filename in uploaded_files:
                        try:
                            src_file = uploaded_files[filename]
                            src_wb = load_workbook(src_file)
                            
                            # Copy each selected sheet
                            for sheet_name in src_wb.sheetnames:
                                if selected_sheets and sheet_name not in selected_sheets:
                                    continue
                                
                                src_ws = src_wb[sheet_name]
                                
                                # Create unique sheet name
                                dest_sheet_name = f"{site.get('name', 'Site')} - {sheet_name}"[:31]
                                if dest_sheet_name in wb.sheetnames:
                                    dest_sheet_name = f"{dest_sheet_name[:28]}_{len(wb.sheetnames)}"
                                
                                dest_ws = wb.create_sheet(title=dest_sheet_name)
                                
                                # Copy dimensions
                                for col_letter, col_dim in src_ws.column_dimensions.items():
                                    dest_ws.column_dimensions[col_letter].width = col_dim.width
                                    dest_ws.column_dimensions[col_letter].hidden = col_dim.hidden
                                
                                for row_num, row_dim in src_ws.row_dimensions.items():
                                    dest_ws.row_dimensions[row_num].height = row_dim.height
                                    dest_ws.row_dimensions[row_num].hidden = row_dim.hidden
                                
                                # Copy cells with formatting
                                for row in src_ws.iter_rows():
                                    for cell in row:
                                        dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                                        
                                        # Copy font
                                        if cell.font:
                                            dest_cell.font = Font(
                                                name=cell.font.name,
                                                size=cell.font.size,
                                                bold=cell.font.bold,
                                                italic=cell.font.italic,
                                                underline=cell.font.underline,
                                                strike=cell.font.strike,
                                                color=cell.font.color
                                            )
                                        
                                        # Copy fill
                                        if cell.fill and cell.fill.patternType:
                                            dest_cell.fill = PatternFill(
                                                patternType=cell.fill.patternType,
                                                fgColor=cell.fill.fgColor,
                                                bgColor=cell.fill.bgColor
                                            )
                                        
                                        # Copy border
                                        if cell.border:
                                            dest_cell.border = Border(
                                                left=cell.border.left,
                                                right=cell.border.right,
                                                top=cell.border.top,
                                                bottom=cell.border.bottom
                                            )
                                        
                                        # Copy alignment
                                        if cell.alignment:
                                            dest_cell.alignment = Alignment(
                                                horizontal=cell.alignment.horizontal,
                                                vertical=cell.alignment.vertical,
                                                wrap_text=cell.alignment.wrap_text,
                                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                                indent=cell.alignment.indent
                                            )
                                        
                                        # Copy number format
                                        if cell.number_format:
                                            dest_cell.number_format = cell.number_format
                                
                                # Copy merged cells
                                for merged_range in src_ws.merged_cells.ranges:
                                    dest_ws.merge_cells(str(merged_range))
                            
                            src_wb.close()
                            
                        except Exception as e:
                            # Create error sheet
                            error_ws = wb.create_sheet(title=f"Erreur - {site.get('name', 'Site')}"[:31])
                            error_ws.cell(row=1, column=1, value=f"Erreur: {str(e)}")
            
            # Ensure at least one sheet exists
            if len(wb.sheetnames) == 0:
                wb.create_sheet(title="Vide")
                wb.active.cell(row=1, column=1, value="Aucune donnée")
            
            # Save
            temp_dir = Path(settings.MEDIA_ROOT) / 'temp'
            temp_dir.mkdir(exist_ok=True)
            safe_filename = "".join(c for c in output_filename if c.isalnum() or c in (' ', '-', '_')).strip()
            output_path = temp_dir / f'{safe_filename}.xlsx'
            wb.save(output_path)
            wb.close()
            
            return FileResponse(
                open(output_path, 'rb'),
                as_attachment=True,
                filename=f'{safe_filename}.xlsx'
            )
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
