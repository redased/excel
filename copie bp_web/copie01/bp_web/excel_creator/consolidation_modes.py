"""
Consolidation Modes - Multiple types of Excel consolidation
Modes:
1. Simple - Stack tables one below another
2. Synthesis - Comparison and summary between sites  
3. Statistics - Calculations (sum, avg, min, max, stdev)
4. Graphs - Charts and visualizations
5. Complete - All of the above
"""
import json
from pathlib import Path
from django.http import JsonResponse, FileResponse
from django.views import View
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


# ============================================
# CONSOLIDATION MODES
# ============================================

CONSOLIDATION_MODES = {
    'simple': {
        'name': 'Simple (Empilage)',
        'description': 'Empiler les tableaux un en dessous de l\'autre',
        'icon': '📋'
    },
    'synthesis': {
        'name': 'Avec Synthèse',
        'description': 'Comparaison et récapitulatif entre sites',
        'icon': '📊'
    },
    'statistics': {
        'name': 'Avec Statistiques',
        'description': 'Calculs: Somme, Moyenne, Min, Max, Écart-type',
        'icon': '📈'
    },
    'graphs': {
        'name': 'Avec Graphiques',
        'description': 'Graphiques de comparaison et d\'évolution',
        'icon': '📉'
    },
    'complete': {
        'name': 'Complet',
        'description': 'Synthèse + Statistiques + Graphiques',
        'icon': '🎯'
    }
}


@method_decorator(csrf_exempt, name='dispatch')
class ConsolidationModesAPIView(View):
    """Return available consolidation modes"""
    
    def get(self, request):
        return JsonResponse({
            'success': True,
            'modes': CONSOLIDATION_MODES
        })


@method_decorator(csrf_exempt, name='dispatch')
class MultiModeConsolidationAPIView(View):
    """Main consolidation API with mode selection"""
    
    def post(self, request):
        try:
            # Get parameters
            mode = request.POST.get('mode', 'simple')
            output_filename = request.POST.get('output_filename', 'Consolidation')
            sheet_name = request.POST.get('sheet_name', '')  # Empty = all sheets
            files = request.FILES.getlist('files')
            
            if not files:
                return JsonResponse({'success': False, 'error': 'Aucun fichier fourni'}, status=400)
            
            # Create output workbook
            wb = Workbook()
            default_sheet = wb.active
            
            # Parse options
            options_str = request.POST.get('options', '{}')
            try:
                options = json.loads(options_str)
            except:
                options = {}
            
            # Dispatch to appropriate handler
            if mode == 'simple':
                self._consolidate_simple(wb, files, sheet_name)
            elif mode == 'synthesis':
                self._consolidate_synthesis(wb, files, sheet_name, options.get('synthesis', {}))
            elif mode == 'statistics':
                self._consolidate_statistics(wb, files, sheet_name, options.get('stats', {}))
            elif mode == 'graphs':
                self._consolidate_graphs(wb, files, sheet_name, options.get('charts', {}))
            elif mode == 'complete':
                # Get complete options which includes subsections
                comp_opts = options.get('complete', {})
                self._consolidate_complete(wb, files, sheet_name, options)
            else:
                self._consolidate_simple(wb, files, sheet_name)
            
            # Remove default sheet if others were created
            if len(wb.sheetnames) > 1 and 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Save
            temp_dir = Path(settings.MEDIA_ROOT) / 'temp'
            temp_dir.mkdir(exist_ok=True)
            safe_filename = "".join(c for c in output_filename if c.isalnum() or c in (' ', '-', '_')).strip()
            output_path = temp_dir / f'{safe_filename}.xlsx'
            wb.save(output_path)
            wb.close()
            
            return FileResponse(
                open(output_path, 'rb'),
                as_attachment=True,
                filename=f'{safe_filename}.xlsx'
            )
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    # =====================================
    # MODE 1: SIMPLE (Empilage)
    # =====================================
    
    def _consolidate_simple(self, wb, files, target_sheet=''):
        """Stack tables one below another"""
        
        # Styles
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        separator_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get unique sheets across all files
        all_sheets = set()
        for f in files:
            try:
                src_wb = load_workbook(f, read_only=True)
                all_sheets.update(src_wb.sheetnames)
                src_wb.close()
                f.seek(0)  # Reset file pointer
            except:
                pass
        
        # Filter sheets if target specified
        if target_sheet:
            sheets_to_process = [target_sheet] if target_sheet in all_sheets else list(all_sheets)[:1]
        else:
            sheets_to_process = list(all_sheets)
        
        # Process each sheet
        for sheet_name in sheets_to_process:
            # Create or get destination sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name[:31])
            
            current_row = 1
            max_cols = 1
            
            # Process each file
            for f in files:
                # Skip temporary Excel files (created when file is open)
                if f.name.startswith('~$'):
                    continue
                    
                try:
                    src_wb = load_workbook(f, data_only=True)
                    
                    if sheet_name not in src_wb.sheetnames:
                        src_wb.close()
                        f.seek(0)
                        continue
                    
                    src_ws = src_wb[sheet_name]
                    
                    # Add file separator/header
                    ws.merge_cells(start_row=current_row, start_column=1, 
                                  end_row=current_row, end_column=max(10, src_ws.max_column))
                    cell = ws.cell(row=current_row, column=1, value=f"📄 {f.name}")
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    current_row += 1
                    
                    # Copy data with formatting
                    for row_idx, row in enumerate(src_ws.iter_rows(min_row=1, max_row=src_ws.max_row), start=1):
                        for col_idx, cell in enumerate(row, start=1):
                            dest_cell = ws.cell(row=current_row, column=col_idx, value=cell.value)
                            
                            # Copy formatting
                            if cell.font:
                                dest_cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    color=cell.font.color
                                )
                            if cell.fill and cell.fill.patternType:
                                try:
                                    dest_cell.fill = PatternFill(
                                        patternType=cell.fill.patternType,
                                        fgColor=cell.fill.fgColor.rgb if cell.fill.fgColor else None,
                                        bgColor=cell.fill.bgColor.rgb if cell.fill.bgColor else None
                                    )
                                except:
                                    pass
                            if cell.border:
                                try:
                                    dest_cell.border = Border(
                                        left=Side(style=cell.border.left.style, color=cell.border.left.color.rgb if cell.border.left.color else None) if cell.border.left else None,
                                        right=Side(style=cell.border.right.style, color=cell.border.right.color.rgb if cell.border.right.color else None) if cell.border.right else None,
                                        top=Side(style=cell.border.top.style, color=cell.border.top.color.rgb if cell.border.top.color else None) if cell.border.top else None,
                                        bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color.rgb if cell.border.bottom.color else None) if cell.border.bottom else None
                                    )
                                except:
                                    pass
                            if cell.alignment:
                                try:
                                    dest_cell.alignment = Alignment(
                                        horizontal=cell.alignment.horizontal,
                                        vertical=cell.alignment.vertical,
                                        wrap_text=cell.alignment.wrap_text,
                                        shrink_to_fit=cell.alignment.shrink_to_fit
                                    )
                                except:
                                    pass
                            if cell.number_format:
                                dest_cell.number_format = cell.number_format
                            
                            max_cols = max(max_cols, col_idx)
                        
                        current_row += 1
                    
                    # Auto-fit column widths based on content
                    for col_idx in range(1, max_cols + 1):
                        max_length = 0
                        col_letter = get_column_letter(col_idx)
                        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                            for cell in row:
                                if cell.value:
                                    cell_length = len(str(cell.value))
                                    max_length = max(max_length, cell_length)
                        # Set width with some padding
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        if adjusted_width > 8:  # Only adjust if content is significant
                            ws.column_dimensions[col_letter].width = adjusted_width
                    
                    # Add spacing between files
                    current_row += 2
                    
                    src_wb.close()
                    f.seek(0)
                    
                except Exception as e:
                    ws.cell(row=current_row, column=1, value=f"Erreur: {f.name} - {str(e)}")
                    current_row += 2
                    current_row += 2
    
    # =====================================
    # MODE 2: SYNTHESIS (Comparaison)
    # =====================================
    
    def _consolidate_synthesis(self, wb, files, target_sheet='', options=None):
        """Create summary and comparison between sites"""
        if options is None: options = {}
        
        # Default to True if no options provided or specific key missing
        show_recap = options.get('recap', True) if options else True
        show_compare = options.get('compare', True) if options else True
        # show_diff = options.get('diff', False) 
        # show_rank = options.get('rank', False)
        
        # First, do simple consolidation
        self._consolidate_simple(wb, files, target_sheet)
        
        # Create synthesis sheet
        synth_ws = wb.create_sheet(title="📊 Synthèse", index=0)
        
        # Styles
        title_font = Font(bold=True, size=14, color="1E293B")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        synth_ws.cell(row=1, column=1, value="📊 Synthèse de Consolidation").font = title_font
        synth_ws.merge_cells('A1:E1')
        
        # Summary table
        synth_ws.cell(row=3, column=1, value="Fichier").fill = header_fill
        synth_ws.cell(row=3, column=1).font = header_font
        synth_ws.cell(row=3, column=2, value="Feuilles").fill = header_fill
        synth_ws.cell(row=3, column=2).font = header_font
        synth_ws.cell(row=3, column=3, value="Lignes").fill = header_fill
        synth_ws.cell(row=3, column=3).font = header_font
        synth_ws.cell(row=3, column=4, value="Colonnes").fill = header_fill
        synth_ws.cell(row=3, column=4).font = header_font
        
        row = 4
        total_rows = 0
        for f in files:
            try:
                src_wb = load_workbook(f, read_only=True)
                sheets = len(src_wb.sheetnames)
                rows = sum(ws.max_row for ws in src_wb.worksheets)
                cols = max(ws.max_column for ws in src_wb.worksheets)
                
                synth_ws.cell(row=row, column=1, value=f.name)
                synth_ws.cell(row=row, column=2, value=sheets)
                synth_ws.cell(row=row, column=3, value=rows)
                synth_ws.cell(row=row, column=4, value=cols)
                
                total_rows += rows
                row += 1
                
                src_wb.close()
                f.seek(0)
            except:
                f.seek(0)
        
        # Total
        synth_ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        synth_ws.cell(row=row, column=2, value=len(files)).font = Font(bold=True)
        synth_ws.cell(row=row, column=3, value=total_rows).font = Font(bold=True)
        
        # Auto-fit columns
        for col in range(1, 5):
            synth_ws.column_dimensions[get_column_letter(col)].width = 20
    
    # =====================================
    # MODE 3: STATISTICS
    # =====================================
    
    def _consolidate_statistics(self, wb, files, target_sheet='', options=None):
        """Add statistical calculations"""
        if options is None: options = {}
        
        # Options
        show_sum = options.get('sum', True)
        show_avg = options.get('avg', True)
        show_min = options.get('min', True)
        show_max = options.get('max', True)
        show_count = options.get('count', False)
        show_stdev = options.get('stdev', False)
        show_median = options.get('median', False)
        show_mode = options.get('mode', False)
        show_var = options.get('var', False)
        show_range = options.get('range', False)
        show_q1 = options.get('q1', False)
        show_q2 = options.get('q2', False)
        show_q3 = options.get('q3', False)
        show_iqr = options.get('iqr', False)
        show_skew = options.get('skew', False)
        show_kurt = options.get('kurt', False)
        show_cv = options.get('cv', False)
        
        # First, do synthesis
        # Pass synthesis options if available in options dict (for complete mode)
        synth_opts = options.get('synthesis', {}) if 'synthesis' in options else {}
        self._consolidate_synthesis(wb, files, target_sheet, synth_opts)
        
        # Create stats sheet
        stats_ws = wb.create_sheet(title="📈 Statistiques", index=1)
        
        # Styles
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        stats_ws.cell(row=1, column=1, value="📈 Statistiques par Fichier").font = Font(bold=True, size=14)
        
        # Build Headers based on options
        headers = ['Fichier']
        if show_sum: headers.append('Somme')
        if show_avg: headers.append('Moyenne')
        if show_median: headers.append('Médiane')
        if show_mode: headers.append('Mode')
        if show_min: headers.append('Min')
        if show_max: headers.append('Max')
        if show_range: headers.append('Étendue')
        if show_count: headers.append('Nombre')
        if show_stdev: headers.append('Écart-type')
        if show_var: headers.append('Variance')
        if show_cv: headers.append('CV %')
        if show_q1: headers.append('Q1 (25%)')
        if show_q2: headers.append('Q2 (50%)')
        if show_q3: headers.append('Q3 (75%)')
        if show_iqr: headers.append('IQR')
        if show_skew: headers.append('Asymétrie')
        if show_kurt: headers.append('Aplatissement')
        
        for col, header in enumerate(headers, start=1):
            cell = stats_ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row = 4
        import statistics
        import math
        
        for f in files:
            try:
                src_wb = load_workbook(f, data_only=True)
                
                # Collect all numeric values
                all_values = []
                for ws in src_wb.worksheets:
                    for row_data in ws.iter_rows():
                        for cell in row_data:
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                all_values.append(cell.value)
                
                if all_values:
                    col = 1
                    stats_ws.cell(row=row, column=col, value=f.name); col += 1
                    
                    if show_sum: stats_ws.cell(row=row, column=col, value=sum(all_values)); col += 1
                    if show_avg: stats_ws.cell(row=row, column=col, value=statistics.mean(all_values)); col += 1
                    if show_median: stats_ws.cell(row=row, column=col, value=statistics.median(all_values)); col += 1
                    if show_mode: 
                        try: stats_ws.cell(row=row, column=col, value=statistics.mode(all_values))
                        except: stats_ws.cell(row=row, column=col, value="N/A")
                        col += 1
                    if show_min: stats_ws.cell(row=row, column=col, value=min(all_values)); col += 1
                    if show_max: stats_ws.cell(row=row, column=col, value=max(all_values)); col += 1
                    if show_range: stats_ws.cell(row=row, column=col, value=max(all_values) - min(all_values)); col += 1
                    if show_count: stats_ws.cell(row=row, column=col, value=len(all_values)); col += 1
                    
                    stdev_val = 0
                    if len(all_values) > 1:
                        stdev_val = statistics.stdev(all_values)
                    
                    if show_stdev: stats_ws.cell(row=row, column=col, value=stdev_val); col += 1
                    if show_var: stats_ws.cell(row=row, column=col, value=statistics.variance(all_values) if len(all_values) > 1 else 0); col += 1
                    if show_cv: 
                        mean_val = statistics.mean(all_values)
                        cv = (stdev_val / mean_val * 100) if mean_val != 0 else 0
                        stats_ws.cell(row=row, column=col, value=cv); col += 1
                    
                    if show_q1 or show_q2 or show_q3 or show_iqr:
                        sorted_vals = sorted(all_values)
                        n = len(sorted_vals)
                        def get_percentile(p):
                            k = (n - 1) * p
                            f = math.floor(k)
                            c = math.ceil(k)
                            if f == c: return sorted_vals[int(k)]
                            return sorted_vals[int(f)] * (c - k) + sorted_vals[int(c)] * (k - f)
                        
                        q1 = get_percentile(0.25)
                        q3 = get_percentile(0.75)
                        
                        if show_q1: stats_ws.cell(row=row, column=col, value=q1); col += 1
                        if show_q2: stats_ws.cell(row=row, column=col, value=statistics.median(all_values)); col += 1
                        if show_q3: stats_ws.cell(row=row, column=col, value=q3); col += 1
                        if show_iqr: stats_ws.cell(row=row, column=col, value=q3 - q1); col += 1

                    # Helper for moments
                    if (show_skew or show_kurt) and len(all_values) > 1:
                         mean = statistics.mean(all_values)
                         std = statistics.stdev(all_values)
                         if std > 0:
                             z_scores = [(x - mean) / std for x in all_values]
                             if show_skew: 
                                 skew = sum(z**3 for z in z_scores) / n
                                 stats_ws.cell(row=row, column=col, value=skew); col += 1
                             if show_kurt:
                                 kurt = sum(z**4 for z in z_scores) / n - 3
                                 stats_ws.cell(row=row, column=col, value=kurt); col += 1
                         else:
                             if show_skew: stats_ws.cell(row=row, column=col, value=0); col += 1
                             if show_kurt: stats_ws.cell(row=row, column=col, value=0); col += 1

                    row += 1
                
                src_wb.close()
                f.seek(0)
            except:
                f.seek(0)
        
        # Format numbers
        for r in range(4, row):
            for c in range(2, len(headers) + 1):
                cell = stats_ws.cell(row=r, column=c)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Auto-fit
        for col in range(1, len(headers) + 1):
            stats_ws.column_dimensions[get_column_letter(col)].width = 18
    
    # =====================================
    # MODE 4: GRAPHS
    # =====================================
    
    def _consolidate_graphs(self, wb, files, target_sheet='', options=None):
        """Add charts and visualizations"""
        if options is None: options = {}
        
        show_bar = options.get('bar', True)
        show_line = options.get('line', True)
        show_pie = options.get('pie', False)
        show_area = options.get('area', False)
        
        # First, do statistics (we need data for charts)
        # We reuse statistics logic but potentially with minimal options if not desired
        # For simplicity, we assume statistics standard options are needed to get the data sheet
        # Or better: we rely on the specific 'statistics' sheet being present.
        # But if this is standalone 'graphs' mode, we must generate stats first.
        
        # Check if 'Statistiques' exists, if not generate it
        if "📈 Statistiques" not in wb.sheetnames:
             # Basic stats for charting
             self._consolidate_statistics(wb, files, target_sheet, {'sum': True, 'avg': True, 'min': True, 'max': True})

        # Create charts sheet
        charts_ws = wb.create_sheet(title="📉 Graphiques", index=2)
        charts_ws.cell(row=1, column=1, value="📉 Visualisations").font = Font(bold=True, size=14)
        
        # Get stats data for charts
        stats_ws = wb["📈 Statistiques"]
        
        # Count data rows
        data_rows = 0
        for row in range(4, stats_ws.max_row + 1):
            if stats_ws.cell(row=row, column=1).value:
                data_rows += 1
        
        if data_rows > 0:
            current_row = 3
            cats = Reference(stats_ws, min_col=1, min_row=4, max_row=3 + data_rows)
            
            if show_bar:
                # Bar chart for sums (Column 2 usually)
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = "Comparaison des Sommes"
                bar_chart.y_axis.title = "Valeur"
                
                data = Reference(stats_ws, min_col=2, min_row=3, max_row=3 + data_rows, max_col=2)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(cats)
                bar_chart.shape = 4
                
                charts_ws.add_chart(bar_chart, f"A{current_row}")
                current_row += 15
            
            if show_line:
                # Line chart for comparison (Min/Max/Avg - Cols 3,4,5 usually)
                line_chart = LineChart()
                line_chart.title = "Évolution Moyenne/Min/Max"
                line_chart.style = 10
                line_chart.y_axis.title = "Valeur"
                
                # Assuming standard columns order: Sum(2), Avg(3), Min(4), Max(5)
                # We try to be safe
                data2 = Reference(stats_ws, min_col=3, min_row=3, max_row=3 + data_rows, max_col=5)
                line_chart.add_data(data2, titles_from_data=True)
                line_chart.set_categories(cats)
                
                charts_ws.add_chart(line_chart, f"A{current_row}")
                current_row += 15
                
            if show_pie and data_rows <= 10: # Pie chart only if few items
                pie_chart = PieChart()
                pie_chart.title = "Répartition (Somme)"
                data = Reference(stats_ws, min_col=2, min_row=3, max_row=3 + data_rows) # Sums
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(cats)
                
                charts_ws.add_chart(pie_chart, f"A{current_row}")
                current_row += 15
                
            if show_area:
                area_chart = AreaChart()
                area_chart.title = "Comparaison d'Aires"
                area_chart.style = 42
                data = Reference(stats_ws, min_col=2, min_row=3, max_row=3 + data_rows, max_col=2)
                area_chart.add_data(data, titles_from_data=True)
                area_chart.set_categories(cats)
                
                charts_ws.add_chart(area_chart, f"A{current_row}")

    # =====================================
    # MODE 5: COMPLETE
    # =====================================
    
    def _consolidate_complete(self, wb, files, target_sheet='', options=None):
        """Complete consolidation with everything"""
        if options is None: options = {}
        comp_opts = options.get('complete', {})
        
        # Always do simple first
        self._consolidate_simple(wb, files, target_sheet)
        
        if comp_opts.get('synthesis', True):
            self._consolidate_synthesis(wb, files, target_sheet, options.get('synthesis', {}))
            
        if comp_opts.get('stats', True):
           self._consolidate_statistics(wb, files, target_sheet, options.get('stats', {}))
           
        if comp_opts.get('charts', True):
           self._consolidate_graphs(wb, files, target_sheet, options.get('charts', {}))
        
        # Add a dashboard sheet using existing stats
        dashboard = wb.create_sheet(title="🎯 Dashboard", index=0)
        
        dashboard.cell(row=1, column=1, value="🎯 Dashboard de Consolidation").font = Font(bold=True, size=16, color="1E293B")
        dashboard.merge_cells('A1:E1')
        
        dashboard.cell(row=3, column=1, value=f"📁 Fichiers traités: {len(files)}").font = Font(size=12)
        dashboard.cell(row=4, column=1, value=f"📋 Feuilles créées: {len(wb.sheetnames)-1}").font = Font(size=12)
        
        dashboard.cell(row=6, column=1, value="Navigation Rapide:").font = Font(bold=True, size=12)
        row = 7
        for sheet_name in wb.sheetnames:
            if sheet_name != "🎯 Dashboard":
                cell = dashboard.cell(row=row, column=1, value=f"  🔗 {sheet_name}")
                cell.hyperlink = f"#'{sheet_name}'!A1"
                cell.font = Font(color="0000FF", underline="single")
                row += 1
        
        dashboard.column_dimensions['A'].width = 50
