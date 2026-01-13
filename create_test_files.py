"""
Script to create test Excel files for Bubble Consolidation testing
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Create folder structure
base_path = r'C:\Users\reda\Desktop\python automatisation\TestResponsables'
os.makedirs(base_path, exist_ok=True)

responsables = {
    'Jean_Dupont': ['Site_Paris', 'Site_Lyon', 'Site_Marseille'],
    'Marie_Martin': ['Site_Nice', 'Site_Lille'],
    'Pierre_Bernard': ['Site_Bordeaux', 'Site_Toulouse', 'Site_Nantes', 'Site_Strasbourg']
}

# Styles
header_fill = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
green_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
orange_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
title_font = Font(bold=True, size=14)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for resp_name, sites in responsables.items():
    resp_folder = os.path.join(base_path, resp_name)
    os.makedirs(resp_folder, exist_ok=True)
    
    for site_name in sites:
        wb = Workbook()
        
        # Sheet 1: Branche
        ws1 = wb.active
        ws1.title = 'Branche'
        
        # Title with merge
        ws1.merge_cells('A1:F1')
        site_display = site_name.replace("_", " ")
        ws1['A1'] = f'Budget Previsionnel 2026 - {site_display}'
        ws1['A1'].font = title_font
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Code', 'Libelle', 'Budget N-1', 'Realise N-1', 'Budget N', 'Ecart']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        data = [
            ['BR001', 'Personnel', 150000, 145000, 160000, 10000],
            ['BR002', 'Fournitures', 25000, 24500, 27000, 2000],
            ['BR003', 'Services', 45000, 43000, 48000, 3000],
            ['BR004', 'Maintenance', 18000, 17500, 19500, 1500],
            ['BR005', 'Formation', 12000, 11000, 14000, 2000],
        ]
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Column widths
        ws1.column_dimensions['A'].width = 10
        ws1.column_dimensions['B'].width = 20
        for col in 'CDEF':
            ws1.column_dimensions[col].width = 15
        
        # Sheet 2: Cout Service
        ws2 = wb.create_sheet('Cout Service')
        
        ws2.merge_cells('A1:E1')
        ws2['A1'] = f'Couts de Service - {site_display}'
        ws2['A1'].font = title_font
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        headers2 = ['Code CS', 'Description', 'Montant', 'Responsable', 'Statut']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.fill = green_fill
            cell.font = header_font
            cell.border = border
        
        resp_display = resp_name.replace('_', ' ')
        data2 = [
            ['CS001', 'Electricite', 8500, resp_display, 'Valide'],
            ['CS002', 'Eau', 2300, resp_display, 'Valide'],
            ['CS003', 'Internet', 1200, resp_display, 'En cours'],
            ['CS004', 'Telephone', 850, resp_display, 'Valide'],
        ]
        for row_idx, row_data in enumerate(data2, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        for col in 'ABCDE':
            ws2.column_dimensions[col].width = 18
        
        # Sheet 3: Recapitulatif
        ws3 = wb.create_sheet('Recapitulatif')
        
        ws3.merge_cells('A1:D1')
        ws3['A1'] = 'Recapitulatif General'
        ws3['A1'].font = Font(bold=True, size=16)
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        ws3['A3'] = 'Site:'
        ws3['B3'] = site_display
        ws3['A4'] = 'Responsable:'
        ws3['B4'] = resp_display
        ws3['A5'] = 'Annee:'
        ws3['B5'] = 2026
        
        ws3.merge_cells('A7:D7')
        ws3['A7'] = 'Totaux par categorie'
        ws3['A7'].font = Font(bold=True, size=12, color='FFFFFF')
        ws3['A7'].fill = orange_fill
        
        ws3['A8'] = 'Total Branche'
        ws3['B8'] = 268500
        ws3['A9'] = 'Total Cout Service'
        ws3['B9'] = 12850
        ws3['A10'] = 'Total General'
        ws3['B10'] = 281350
        ws3['B10'].font = Font(bold=True)
        
        for col in 'ABCD':
            ws3.column_dimensions[col].width = 20
        
        # Save
        filepath = os.path.join(resp_folder, f'{site_name}.xlsx')
        wb.save(filepath)
        print(f'Cree: {filepath}')

print('\n--- Fichiers de test crees dans: TestResponsables/ ---')
print('\nStructure:')
for resp_name, sites in responsables.items():
    print(f'  {resp_name}/')
    for site in sites:
        print(f'    - {site}.xlsx (3 feuilles)')
