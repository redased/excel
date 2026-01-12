"""
Excel file generator using openpyxl
"""
import os
import uuid
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings


def generate_excel(config: dict) -> str:
    """
    Generate Excel file from configuration
    Returns the filename of the generated file
    """
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    sheets = config.get("sheets", [])
    
    for sheet_config in sheets:
        ws = wb.create_sheet(title=sheet_config.get("name", "Feuille1"))
        
        # Apply column configurations
        for col in sheet_config.get("columns", []):
            col_letter = get_column_letter(col["index"])
            ws.column_dimensions[col_letter].width = col.get("width", 12)
            
            if col.get("header"):
                cell = ws.cell(row=1, column=col["index"])
                cell.value = col["header"]
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply row configurations
        for row in sheet_config.get("rows", []):
            ws.row_dimensions[row["index"]].height = row.get("height", 15)
        
        # Apply cell contents
        for cell_config in sheet_config.get("cells", []):
            cell = ws.cell(row=cell_config["row"], column=cell_config["col"])
            cell.value = cell_config.get("value", "")
            
            style = cell_config.get("style", {})
            
            # Font
            cell.font = Font(
                size=style.get("font_size", 11),
                color=style.get("font_color", "000000"),
                bold=style.get("bold", False),
                italic=style.get("italic", False),
                underline='single' if style.get("underline") else None
            )
            
            # Background
            if style.get("bg_color"):
                cell.fill = PatternFill(
                    start_color=style["bg_color"],
                    end_color=style["bg_color"],
                    fill_type="solid"
                )
            
            # Alignment
            cell.alignment = Alignment(
                horizontal=style.get("alignment", "left"),
                vertical=style.get("vertical_alignment", "center"),
                wrap_text=style.get("wrap_text", False)
            )
            
            # Border
            if style.get("border_style"):
                side = Side(style=style["border_style"], color=style.get("border_color", "000000"))
                cell.border = Border(left=side, right=side, top=side, bottom=side)
        
        # Apply merges
        for merge in sheet_config.get("merges", []):
            ws.merge_cells(
                start_row=merge["start_row"],
                start_column=merge["start_col"],
                end_row=merge["end_row"],
                end_column=merge["end_col"]
            )
    
    # Generate unique filename
    filename = f"BP_2026_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = Path(settings.MEDIA_ROOT) / "excel" / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(filepath)
    
    return filename
