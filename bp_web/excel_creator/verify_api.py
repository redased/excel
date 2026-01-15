"""
File Verification API - Compare source files with generated result and detect/correct errors
"""
from django.views import View
from django.http import JsonResponse, FileResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import tempfile
import os


@method_decorator(csrf_exempt, name='dispatch')
class FileCompareAPIView(View):
    """Compare source files with result file and detect errors"""
    
    def post(self, request):
        try:
            source_files = request.FILES.getlist('source_files')
            result_file = request.FILES.get('result_file')
            sheet_name = request.POST.get('sheet_name', '')
            
            if not source_files or not result_file:
                return JsonResponse({'success': False, 'error': 'Fichiers manquants'}, status=400)
            
            # Load result file
            result_wb = load_workbook(result_file, data_only=True)
            result_file.seek(0)
            
            if sheet_name not in result_wb.sheetnames:
                sheet_name = result_wb.sheetnames[0] if result_wb.sheetnames else ''
            
            if not sheet_name:
                return JsonResponse({'success': False, 'error': 'Aucune feuille disponible'}, status=400)
            
            result_ws = result_wb[sheet_name]
            
            # Collect values from source files
            source_data = {}  # {(row, col): [values]}
            source_file_names = []
            
            for f in source_files:
                source_file_names.append(f.name)
                try:
                    src_wb = load_workbook(f, data_only=True)
                    if sheet_name in src_wb.sheetnames:
                        src_ws = src_wb[sheet_name]
                        for row_idx in range(1, src_ws.max_row + 1):
                            for col_idx in range(1, src_ws.max_column + 1):
                                cell = src_ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None and isinstance(cell.value, (int, float)):
                                    key = (row_idx, col_idx)
                                    if key not in source_data:
                                        source_data[key] = []
                                    source_data[key].append(cell.value)
                    src_wb.close()
                except Exception as e:
                    print(f"Error loading {f.name}: {e}")
                finally:
                    f.seek(0)
            
            # Compare with result
            comparison_table = []
            errors = []
            correct_count = 0
            error_count = 0
            
            for (row_idx, col_idx), source_values in source_data.items():
                if not source_values:
                    continue
                
                expected_sum = sum(source_values)
                result_cell = result_ws.cell(row=row_idx, column=col_idx)
                actual_value = result_cell.value
                
                # Skip if result is None or not numeric
                if actual_value is None or not isinstance(actual_value, (int, float)):
                    continue
                
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                difference = abs(expected_sum - actual_value)
                
                # Consider equal if difference is very small (floating point tolerance)
                is_correct = difference < 0.01
                
                status = 'ok' if is_correct else 'error'
                
                row_data = {
                    'cell': cell_ref,
                    'row': row_idx,
                    'col': col_idx,
                    'source_values': source_values,
                    'result_value': actual_value,
                    'expected_value': expected_sum,
                    'status': status
                }
                
                comparison_table.append(row_data)
                
                if is_correct:
                    correct_count += 1
                else:
                    error_count += 1
                    errors.append({
                        'cell': cell_ref,
                        'row': row_idx,
                        'col': col_idx,
                        'actual': actual_value,
                        'expected': expected_sum,
                        'difference': expected_sum - actual_value
                    })
            
            # Sort by row then column
            comparison_table.sort(key=lambda x: (x['row'], x['col']))
            
            result_wb.close()
            
            return JsonResponse({
                'success': True,
                'source_file_names': source_file_names,
                'sheet_name': sheet_name,
                'comparison_table': comparison_table[:100],  # Limit to 100 rows
                'errors': errors,
                'correct_count': correct_count,
                'error_count': error_count
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class FileCorrectAPIView(View):
    """Correct errors in result file"""
    
    def post(self, request):
        try:
            result_file = request.FILES.get('result_file')
            sheet_name = request.POST.get('sheet_name', '')
            corrections_json = request.POST.get('corrections', '[]')
            
            if not result_file:
                return JsonResponse({'success': False, 'error': 'Fichier résultat manquant'}, status=400)
            
            corrections = json.loads(corrections_json)
            
            # Load result file (with formulas preserved)
            wb = load_workbook(result_file, data_only=False)
            
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0] if wb.sheetnames else ''
            
            if not sheet_name:
                return JsonResponse({'success': False, 'error': 'Aucune feuille disponible'}, status=400)
            
            ws = wb[sheet_name]
            
            # Apply corrections
            for correction in corrections:
                cell_ref = correction.get('cell', '')
                value = correction.get('value', 0)
                
                if cell_ref:
                    ws[cell_ref] = value
            
            # Save to temp file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            wb.close()
            temp_file.close()
            
            # Return file
            response = FileResponse(
                open(temp_file.name, 'rb'),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                filename='Consolidation_Corrigee.xlsx'
            )
            
            # Schedule cleanup after response
            response._temp_file_path = temp_file.name
            
            return response
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class FilePreviewAPIView(View):
    """Read Excel file content for preview"""
    
    def post(self, request):
        try:
            file = request.FILES.get('file')
            sheet_name = request.POST.get('sheet_name', '')
            
            if not file:
                return JsonResponse({'success': False, 'error': 'Fichier manquant'}, status=400)
            
            # Load workbook
            wb = load_workbook(file, data_only=True)
            
            # Get sheet names
            sheet_names = wb.sheetnames
            
            # Use first sheet if not specified
            if not sheet_name or sheet_name not in sheet_names:
                sheet_name = sheet_names[0] if sheet_names else ''
            
            if not sheet_name:
                return JsonResponse({'success': False, 'error': 'Aucune feuille disponible'}, status=400)
            
            ws = wb[sheet_name]
            
            # Read data (limit to 100 rows, 20 columns for preview)
            max_row = min(ws.max_row or 1, 100)
            max_col = min(ws.max_column or 1, 20)
            
            rows = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    # Format value for display
                    if value is None:
                        display = ''
                    elif isinstance(value, (int, float)):
                        display = f'{value:,.2f}' if isinstance(value, float) and value != int(value) else str(int(value) if isinstance(value, float) else value)
                    else:
                        display = str(value)[:50]  # Truncate long text
                    row_data.append(display)
                rows.append(row_data)
            
            # Get column headers (A, B, C, ...)
            headers = [get_column_letter(i) for i in range(1, max_col + 1)]
            
            wb.close()
            
            return JsonResponse({
                'success': True,
                'sheet_names': sheet_names,
                'current_sheet': sheet_name,
                'headers': headers,
                'rows': rows,
                'total_rows': ws.max_row or 0,
                'total_cols': ws.max_column or 0
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

