"""
Views for Excel Creator app
"""
import json
import os
from pathlib import Path

from django.shortcuts import render
from django.http import JsonResponse, FileResponse, Http404
from django.views import View
from django.views.generic import TemplateView
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import load_workbook

from .models import ExcelConfiguration
from .ai_generator import generate_config_with_ai
from .excel_generator import generate_excel


class IndexView(TemplateView):
    """Main page view"""
    template_name = 'excel_creator/index.html'


@method_decorator(csrf_exempt, name='dispatch')
class GenerateExcelAPIView(View):
    """API endpoint to generate Excel from configuration"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            config = data.get('config', {})
            
            if not config.get('sheets'):
                return JsonResponse({
                    'success': False,
                    'error': 'Configuration invalide: aucune feuille définie'
                }, status=400)
            
            # Generate Excel file
            filename = generate_excel(config)
            
            # Save configuration to history
            ExcelConfiguration.objects.create(
                name=data.get('name', 'BP 2026'),
                config_json=config
            )
            
            return JsonResponse({
                'success': True,
                'filename': filename,
                'download_url': f'/api/download/{filename}/'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'JSON invalide'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class AIGenerateAPIView(View):
    """API endpoint to generate configuration using AI"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            prompt = data.get('prompt', '')
            api_key = data.get('api_key', '')
            
            if not prompt:
                return JsonResponse({
                    'success': False,
                    'error': 'Veuillez entrer une description'
                }, status=400)
            
            if not api_key:
                return JsonResponse({
                    'success': False,
                    'error': 'Veuillez entrer votre clé API Z.ai'
                }, status=400)
            
            # Generate configuration using AI
            config = generate_config_with_ai(prompt, api_key)
            
            # Save to history
            ExcelConfiguration.objects.create(
                name='BP 2026 (IA)',
                config_json=config,
                prompt=prompt
            )
            
            return JsonResponse({
                'success': True,
                'config': config
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'JSON invalide'
            }, status=400)
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur: {str(e)}'
            }, status=500)


class DownloadExcelView(View):
    """Download generated Excel file"""
    
    def get(self, request, filename):
        filepath = Path(settings.MEDIA_ROOT) / "excel" / filename
        
        if not filepath.exists():
            raise Http404("Fichier non trouvé")
        
        response = FileResponse(
            open(filepath, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


@method_decorator(csrf_exempt, name='dispatch')
class UploadFilesAPIView(View):
    """API endpoint to upload Excel files for consolidation"""
    
    def post(self, request):
        try:
            files = request.FILES.getlist('files')
            
            if not files:
                return JsonResponse({
                    'success': False,
                    'error': 'Aucun fichier uploadé'
                }, status=400)
            
            uploaded = []
            upload_dir = Path(settings.MEDIA_ROOT) / "uploads"
            upload_dir.mkdir(parents=True, exist_ok=True)
            
            for f in files:
                filepath = upload_dir / f.name
                with open(filepath, 'wb+') as dest:
                    for chunk in f.chunks():
                        dest.write(chunk)
                uploaded.append({
                    'name': f.name,
                    'path': str(filepath),
                    'sheets': []
                })
                
                # Try to read sheets
                try:
                    wb = load_workbook(filepath, read_only=True, keep_links=False)
                    uploaded[-1]['sheets'] = wb.sheetnames
                    wb.close()
                except Exception:
                    pass
            
            return JsonResponse({
                'success': True,
                'files': uploaded
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class ConsolidateAPIView(View):
    """API endpoint to consolidate Excel files"""
    
    def post(self, request):
        try:
            from .consolidator import consolidate_files
            
            data = json.loads(request.body)
            files = data.get('files', [])
            
            if not files:
                return JsonResponse({
                    'success': False,
                    'error': 'Aucun fichier à consolider'
                }, status=400)
            
            filename = consolidate_files(
                files=files,
                sheet_name=data.get('sheet_name', 'Feuille1'),
                start_column=data.get('start_column', 'E'),
                end_column=data.get('end_column', 'P'),
                start_row=data.get('start_row', 1),
                end_row=data.get('end_row', 10),
                group_by=data.get('group_by', 'branch')
            )
            
            return JsonResponse({
                'success': True,
                'filename': filename,
                'download_url': f'/api/download/{filename}/'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class AIConsolidateAPIView(View):
    """API endpoint to analyze files with AI for consolidation"""
    
    def post(self, request):
        try:
            from .consolidator import analyze_with_ai
            
            data = json.loads(request.body)
            files = data.get('files', [])
            description = data.get('description', '')
            api_key = data.get('api_key', '')
            
            if not api_key:
                return JsonResponse({
                    'success': False,
                    'error': 'Clé API requise'
                }, status=400)
            
            config = analyze_with_ai(files, description, api_key)
            
            return JsonResponse({
                'success': True,
                'config': config
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class ParseBubbleFilesAPIView(View):
    """API endpoint to parse uploaded files for bubble visualization"""
    
    def post(self, request):
        try:
            files = request.FILES.getlist('files')
            paths = request.POST.getlist('paths')
            
            if not files:
                return JsonResponse({
                    'success': False,
                    'error': 'Aucun fichier uploadé'
                }, status=400)
            
            # Group files by responsable (folder name)
            responsables = {}
            
            for i, file in enumerate(files):
                path = paths[i] if i < len(paths) else file.name
                path_parts = path.replace('\\', '/').split('/')
                
                # Extract responsable name from path
                if len(path_parts) >= 2:
                    responsable_name = path_parts[-2].replace('_', ' ')
                else:
                    responsable_name = 'Responsable'
                
                if responsable_name not in responsables:
                    responsables[responsable_name] = {
                        'id': f'resp_{len(responsables)}',
                        'name': responsable_name,
                        'sites': []
                    }
                
                # Parse Excel file to get sheets
                try:
                    wb = load_workbook(file, read_only=True, data_only=True)
                    sheets = []
                    for idx, sheet_name in enumerate(wb.sheetnames):
                        ws = wb[sheet_name]
                        # Get column letters
                        columns = []
                        if ws.max_column:
                            for col_idx in range(1, min(ws.max_column + 1, 27)):
                                from openpyxl.utils import get_column_letter
                                columns.append(get_column_letter(col_idx))
                        
                        sheets.append({
                            'id': f'sh_{len(responsables)}_{len(responsables[responsable_name]["sites"])}_{idx}',
                            'name': sheet_name,
                            'columns': columns
                        })
                    wb.close()
                except Exception as e:
                    sheets = [
                        {'id': f'sh_default_0', 'name': 'Feuille1', 'columns': ['A', 'B', 'C', 'D', 'E']}
                    ]
                
                site_name = file.name.replace('.xlsx', '').replace('.xls', '').replace('_', ' ')
                responsables[responsable_name]['sites'].append({
                    'id': f'site_{len(responsables)}_{len(responsables[responsable_name]["sites"])}',
                    'name': site_name,
                    'filename': file.name,
                    'sheets': sheets
                })
            
            return JsonResponse({
                'responsables': list(responsables.values())
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GenerateBubbleExcelAPIView(View):
    """API endpoint to generate Excel from bubble configuration"""
    
    def post(self, request):
        try:
            import tempfile
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            structure = json.loads(request.POST.get('structure', '[]'))
            bubble_data = json.loads(request.POST.get('bubbleData', '{}'))
            files = request.FILES.getlist('files')
            
            if not structure:
                return JsonResponse({
                    'success': False,
                    'error': 'Aucune structure définie'
                }, status=400)
            
            # Create file mapping
            file_map = {f.name: f for f in files}
            
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Consolidation"
            
            # Header style
            header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center")
            
            current_row = 1
            
            for item in structure:
                if item['type'] == 'responsable':
                    # Add responsable header
                    ws.cell(row=current_row, column=1, value=f"Responsable: {item['info']['name']}")
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                    current_row += 2
                    
                elif item['type'] == 'site':
                    # Add site data
                    ws.cell(row=current_row, column=1, value=f"Site: {item['info']['name']}")
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="059669")
                    current_row += 1
                    
                    # Try to load actual data from file
                    for resp in bubble_data.get('responsables', []):
                        for site in resp.get('sites', []):
                            if site.get('id') == item['id']:
                                filename = site.get('filename')
                                if filename and filename in file_map:
                                    try:
                                        src_wb = load_workbook(file_map[filename])
                                        for sheet in src_wb.sheetnames[:2]:  # First 2 sheets
                                            src_ws = src_wb[sheet]
                                            ws.cell(row=current_row, column=2, value=f"Feuille: {sheet}")
                                            current_row += 1
                                            
                                            # Copy first 10 rows of data
                                            for row_idx in range(1, min(11, src_ws.max_row + 1)):
                                                for col_idx in range(1, min(6, src_ws.max_column + 1)):
                                                    src_cell = src_ws.cell(row=row_idx, column=col_idx)
                                                    ws.cell(row=current_row, column=col_idx + 1, value=src_cell.value)
                                                current_row += 1
                                            current_row += 1
                                        src_wb.close()
                                    except:
                                        pass
                    current_row += 1
                    
                elif item['type'] == 'sheet':
                    # Add sheet header
                    ws.cell(row=current_row, column=1, value=f"Feuille: {item['info']['name']}")
                    ws.cell(row=current_row, column=1).font = Font(bold=True, color="D97706")
                    current_row += 2
            
            # Save to temp file
            temp_dir = Path(settings.MEDIA_ROOT) / 'temp'
            temp_dir.mkdir(exist_ok=True)
            output_path = temp_dir / 'Consolidation_Bulles.xlsx'
            wb.save(output_path)
            wb.close()
            
            return FileResponse(
                open(output_path, 'rb'),
                as_attachment=True,
                filename='Consolidation_Bulles.xlsx'
            )
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

