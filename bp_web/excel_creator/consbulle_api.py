"""
API Views for Consolidation par Bulle with SQLite persistence
"""
import json
import os
from pathlib import Path

from django.http import JsonResponse, FileResponse
from django.views import View
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.numbers import FORMAT_NUMBER_00

from .models import ConsolidationConfig, Responsable, Site, SiteSheetConfig


# ============================================
# CONFIG CRUD APIs
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class ConsolidationConfigListAPIView(View):
    """List and create consolidation configs"""
    
    def get(self, request):
        """List all saved configurations"""
        configs = ConsolidationConfig.objects.all()
        return JsonResponse({
            'success': True,
            'configs': [
                {
                    'id': c.id,
                    'name': c.name,
                    'output_filename': c.output_filename,
                    'updated_at': c.updated_at.isoformat(),
                    'responsables_count': c.responsables.count()
                }
                for c in configs
            ]
        })
    
    def post(self, request):
        """Create a new configuration"""
        try:
            data = json.loads(request.body)
            config = ConsolidationConfig.objects.create(
                name=data.get('name', 'Nouvelle Configuration'),
                output_filename=data.get('output_filename', 'Consolidation'),
                selected_sheets=data.get('selected_sheets', [])
            )
            
            # Save nested responsables
            if 'responsables' in data:
                for idx, resp_data in enumerate(data['responsables']):
                    resp = Responsable.objects.create(
                        config=config,
                        name=resp_data.get('name', f'Responsable {idx+1}'),
                        order=resp_data.get('order', idx)
                    )
                    
                    if 'sites' in resp_data:
                        for s_idx, site_data in enumerate(resp_data['sites']):
                            Site.objects.create(
                                responsable=resp,
                                name=site_data.get('name', f'Site {s_idx+1}'),
                                original_filename=site_data.get('original_filename', ''),
                                detected_sheets=site_data.get('detected_sheets', []),
                                order=site_data.get('order', s_idx)
                            )
            
            return JsonResponse({
                'success': True,
                'config': config.to_dict()
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)


@method_decorator(csrf_exempt, name='dispatch')
class ConsolidationConfigDetailAPIView(View):
    """Get, update, delete a specific config"""
    
    def get(self, request, config_id):
        """Get full configuration with all relationships"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            return JsonResponse({
                'success': True,
                'config': config.to_dict()
            })
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)
    
    def put(self, request, config_id):
        """Update configuration"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            data = json.loads(request.body)
            
            if 'name' in data:
                config.name = data['name']
            if 'output_filename' in data:
                config.output_filename = data['output_filename']
            if 'selected_sheets' in data:
                config.selected_sheets = data['selected_sheets']
            
            # Update nested tables if provided
            if 'responsables' in data:
                # Clear existing
                config.responsables.all().delete()
                
                # Recreate
                for idx, resp_data in enumerate(data['responsables']):
                    resp = Responsable.objects.create(
                        config=config,
                        name=resp_data.get('name', f'Responsable {idx+1}'),
                        order=resp_data.get('order', idx)
                    )
                    
                    if 'sites' in resp_data:
                        for s_idx, site_data in enumerate(resp_data['sites']):
                            Site.objects.create(
                                responsable=resp,
                                name=site_data.get('name', f'Site {s_idx+1}'),
                                original_filename=site_data.get('original_filename', ''),
                                detected_sheets=site_data.get('detected_sheets', []),
                                order=site_data.get('order', s_idx)
                            )
            
            config.save()
            return JsonResponse({'success': True, 'config': config.to_dict()})
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    def delete(self, request, config_id):
        """Delete configuration"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            config.delete()
            return JsonResponse({'success': True})
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)


# ============================================
# RESPONSABLE APIs
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class ResponsableAPIView(View):
    """Create, update, delete responsables"""
    
    def post(self, request, config_id):
        """Add a responsable to a config"""
        try:
            config = ConsolidationConfig.objects.get(id=config_id)
            data = json.loads(request.body)
            
            resp = Responsable.objects.create(
                config=config,
                name=data.get('name', 'Nouveau Responsable'),
                order=config.responsables.count()
            )
            return JsonResponse({'success': True, 'responsable': resp.to_dict()})
        except ConsolidationConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Config not found'}, status=404)
    
    def put(self, request, config_id, resp_id):
        """Update a responsable"""
        try:
            resp = Responsable.objects.get(id=resp_id, config_id=config_id)
            data = json.loads(request.body)
            
            if 'name' in data:
                resp.name = data['name']
            if 'order' in data:
                resp.order = data['order']
            
            resp.save()
            return JsonResponse({'success': True, 'responsable': resp.to_dict()})
        except Responsable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Responsable not found'}, status=404)
    
    def delete(self, request, config_id, resp_id):
        """Delete a responsable"""
        try:
            resp = Responsable.objects.get(id=resp_id, config_id=config_id)
            resp.delete()
            return JsonResponse({'success': True})
        except Responsable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Responsable not found'}, status=404)


# ============================================
# SITE APIs
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class SiteAPIView(View):
    """Create, update, delete sites"""
    
    def post(self, request, resp_id):
        """Add a site to a responsable"""
        try:
            resp = Responsable.objects.get(id=resp_id)
            data = json.loads(request.body)
            
            site = Site.objects.create(
                responsable=resp,
                name=data.get('name', 'Nouveau Site'),
                original_filename=data.get('original_filename', ''),
                detected_sheets=data.get('detected_sheets', []),
                order=resp.sites.count()
            )
            return JsonResponse({'success': True, 'site': site.to_dict()})
        except Responsable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Responsable not found'}, status=404)
    
    def put(self, request, resp_id, site_id):
        """Update a site"""
        try:
            site = Site.objects.get(id=site_id, responsable_id=resp_id)
            data = json.loads(request.body)
            
            if 'name' in data:
                site.name = data['name']
            if 'detected_sheets' in data:
                site.detected_sheets = data['detected_sheets']
            if 'order' in data:
                site.order = data['order']
            
            site.save()
            return JsonResponse({'success': True, 'site': site.to_dict()})
        except Site.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Site not found'}, status=404)
    
    def delete(self, request, resp_id, site_id):
        """Delete a site"""
        try:
            site = Site.objects.get(id=site_id, responsable_id=resp_id)
            site.delete()
            return JsonResponse({'success': True})
        except Site.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Site not found'}, status=404)


# ============================================
# PARSE EXCEL SHEETS API
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class ParseExcelSheetsAPIView(View):
    """Parse Excel files and return sheet names"""
    
    def post(self, request):
        """Parse uploaded files and return sheet names"""
        try:
            files = request.FILES.getlist('files')
            all_sheets = set()
            file_sheets = {}
            
            # Ensure upload directory exists
            upload_dir = Path(settings.MEDIA_ROOT) / "uploads"
            upload_dir.mkdir(parents=True, exist_ok=True)
            
            for f in files:
                try:
                    # Save file to disk for Preview access
                    filepath = upload_dir / f.name
                    with open(filepath, 'wb+') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)
                    
                    # Read sheets using read_only mode
                    wb = load_workbook(filepath, read_only=True, keep_links=False)
                    sheets = wb.sheetnames
                    file_sheets[f.name] = sheets
                    all_sheets.update(sheets)
                    wb.close()
                except Exception as e:
                    file_sheets[f.name] = {'error': str(e)}
            
            return JsonResponse({
                'success': True,
                'all_sheets': sorted(list(all_sheets)),
                'file_sheets': file_sheets
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)



@method_decorator(csrf_exempt, name='dispatch')
class AnalyzeStructureAPIView(View):
    """Analyze Excel structure to suggest fixed headers"""
    
    def post(self, request):
        try:
            scan_depth = int(request.POST.get('scan_depth', 50))
            filenames = request.POST.getlist('filenames')
            
            sheet_configs = {} # { filename: { sheetname: { row: 2, col: 'B' } } }
            upload_dir = Path(settings.MEDIA_ROOT) / "uploads"
            
            for filename in filenames:
                filepath = upload_dir / filename
                if not filepath.exists():
                    continue
                    
                try:
                    wb = load_workbook(filepath, read_only=True, data_only=True)
                    sheet_configs[filename] = {}
                    
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        # 1. Detect Fixed Rows (Header)
                        # Heuristic: Find first row where > 50% of cells are numeric
                        fixed_row = 1
                        numeric_count = 0
                        total_cells = 0
                        
                        # Scan rows up to scan_depth
                        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_depth, max_col=20)):
                            r = row_idx + 1
                            row_numerics = 0
                            row_total = 0
                            
                            for cell in row:
                                if cell.value is None: continue
                                row_total += 1
                                if isinstance(cell.value, (int, float)):
                                    row_numerics += 1
                            
                            # If row is mostly numeric (>50%), we found the data start
                            # So fixed header is the previous row
                            if row_total > 0 and (row_numerics / row_total) > 0.5:
                                fixed_row = max(1, r - 1)
                                break
                            
                            # Fallback if no numeric row found deep enough: assume 1
                            fixed_row = r 

                        # 2. Detect Fixed Cols (Key)
                        # Heuristic: Find first col where > 50% of cells are numeric
                        fixed_col_idx = 1
                        
                        # Scan cols up to scan_depth (or 20 max)
                        max_scan_col = min(scan_depth, 20)
                        for col_idx in range(1, max_scan_col + 1):
                            col_numerics = 0
                            col_total = 0
                            col_letter = get_column_letter(col_idx)
                            
                            # Check first 20 rows of this column
                            for row in ws.iter_rows(min_row=1, max_row=20, min_col=col_idx, max_col=col_idx):
                                cell = row[0]
                                if cell.value is None: continue
                                col_total += 1
                                if isinstance(cell.value, (int, float)):
                                    col_numerics += 1
                            
                            # If column is mostly numeric, it's DATA. So fixed cols end at previous col.
                            if col_total > 0 and (col_numerics / col_total) > 0.5:
                                fixed_col_idx = max(1, col_idx - 1)
                                break
                            
                            fixed_col_idx = col_idx

                        sheet_configs[filename][sheet_name] = {
                            'row': fixed_row,
                            'col': get_column_letter(fixed_col_idx)
                        }
                        
                    wb.close()
                except Exception as e:
                    print(f"Error analyzing {filename}: {e}")
            
            return JsonResponse({'success': True, 'sheet_configs': sheet_configs})
            
        except Exception as e:
             return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# GENERATE EXCEL WITH FORMAT PRESERVATION
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class GenerateConsBulleV2APIView(View):
    """Generate consolidated Excel with format preservation"""
    
    def post(self, request):
        try:
            config_str = request.POST.get('config', '{}')
            config_data = json.loads(config_str)
            output_filename = config_data.get('output_filename', 'Consolidation')
            selected_sheets = config_data.get('selected_sheets', [])
            files = request.FILES.getlist('files')
            
            # 1. Identify MASTER TEMPLATE (First File)
            # We assume the first file in the list is the intended template
            if not files:
                 return JsonResponse({'success': False, 'error': 'No files uploaded'}, status=400)
            
            master_file = files[0]
            master_filename = master_file.name
            
            # Helper to find file by name
            def get_file_path(fname):
                 return Path(settings.MEDIA_ROOT) / "uploads" / fname

            # Load Master WB
            # We need one instance to WRITE (wb_out)
            # And we strictly need to read its values for summation too, checking formulas...
            # Actually, we can load it once. content is in access.
            
            master_path = get_file_path(master_filename)
            if not master_path.exists():
                # Fallback to saving it if not yet saved (it was saved in parse, but let's be safe)
                 pass 

            # We use the previously saved path from Parse step, assuming they are in uploads/
            # But here `files` are InMemoryUploadedFile objects.
            # Let's ensure they are on disk to be loaded by openpyxl
            upload_dir = Path(settings.MEDIA_ROOT) / "uploads"
            upload_dir.mkdir(parents=True, exist_ok=True)
            saved_paths = {}
            for f in files:
                f_path = upload_dir / f.name
                if not f_path.exists(): # Should exist from parse step, but if not:
                    with open(f_path, 'wb+') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)
                saved_paths[f.name] = f_path

            # Load Master Workbook as the Output Target
            # This PRESERVES Logic, Macros, Images, Charts of the first file
            wb = load_workbook(saved_paths[master_filename]) 
            
            # Parse per-sheet configs
            sheet_configs = config_data.get('sheet_configs', {})

            # 2. Iterate Master Sheets
            for sheet_name in wb.sheetnames:
                # Check if this sheet is selected for consolidation
                # If selected_sheets is active, and this sheet not in it, we might want to SKIP logic OR DELETE it?
                # User says "feuil a ne pas toucher". So if not selected, we TOUCH NOTHING.
                if selected_sheets and sheet_name not in selected_sheets:
                    continue 

                ws = wb[sheet_name]
                
                # Identify Sources for this sheet
                # We reuse sheets_data logic or just check all files?
                # sheets_data was: { 'SheetName': [ (SiteName, SrcFile), ... ] }
                # Let's rebuild a simple list of source files that HAVE this sheet
                sources_for_sheet = []
                for f in files:
                    if f.name == master_filename: continue # Skip master here, we add it later logic? 
                    # No, we need master values too if we zero out and sum.
                    # Actually, we are MODIFYING master.
                    
                    # Check if file has this sheet
                    # Optimal: Open once. But simple loop is safer for now.
                    try:
                        f_wb = load_workbook(saved_paths[f.name], read_only=True)
                        if sheet_name in f_wb.sheetnames:
                            sources_for_sheet.append(f.name)
                        f_wb.close()
                    except:
                        pass
                
                if not sources_for_sheet:
                    continue # No other files have this sheet, keep master as is
                
                # SUMMATION LOGIC
                # 1. Build Sum Matrix from (Master + All Sources)
                
                sum_matrix = {} # {(row, col): value}
                
                # Config for this sheet
                fix_row_start, fix_row_end = 1, 39
                fix_col_start, fix_col_end = 1, 1 # Col A
                
                if master_filename in sheet_configs and sheet_name in sheet_configs[master_filename]:
                    try:
                        sc = sheet_configs[master_filename][sheet_name]
                        fix_row_start = int(sc.get('row_start', 1))
                        fix_row_end = int(sc.get('row_end', 39))
                        fix_col_start = column_index_from_string(sc.get('col_start', 'A'))
                        fix_col_end = column_index_from_string(sc.get('col_end', 'A'))
                    except:
                        pass

                # Scan Function
                def scan_file_values(filepath):
                    _wb = load_workbook(filepath, data_only=True)
                    if sheet_name in _wb.sheetnames:
                        _ws = _wb[sheet_name]
                        for row in _ws.iter_rows():
                            for cell in row:
                                r, c = cell.row, cell.column
                                # Skip Fixed Zones (Logic optimization)
                                if (fix_row_start <= r <= fix_row_end) or (fix_col_start <= c <= fix_col_end):
                                    continue
                                
                                if isinstance(cell.value, (int, float)):
                                    if (r, c) not in sum_matrix:
                                        sum_matrix[(r, c)] = 0
                                    sum_matrix[(r, c)] += cell.value
                    _wb.close()

                # Scan ALL files including Master (to get its values)
                scan_file_values(saved_paths[master_filename])
                for fname in sources_for_sheet:
                    scan_file_values(saved_paths[fname])
                
                # 2. Update Master Sheet (ws)
                # ws is NOT data_only=True, so it has Formulas. Good.
                for row in ws.iter_rows():
                    for cell in row:
                        r, c = cell.row, cell.column
                         # Skip Fixed Zones
                        if (fix_row_start <= r <= fix_row_end) or (fix_col_start <= c <= fix_col_end):
                            continue
                        
                        # Preserve Percentage / Ratio Formulas
                        is_ratio = isinstance(cell.value, str) and cell.value.startswith('=') and '/' in cell.value
                        is_pct = cell.number_format and '%' in cell.number_format
                        
                        if is_ratio or is_pct:
                            continue
                        
                        # OTHERWISE: OVERWRITE WITH SUM
                        # We want to overwrite if it's a Formula (Sum, etc) OR a Number.
                        # We do NOT want to overwrite Text/Labels.
                        
                        is_formula = isinstance(cell.value, str) and cell.value.startswith('=')
                        is_number = isinstance(cell.value, (int, float))
                        
                        if is_formula or is_number:
                            # Use collected sum, or 0 if nothing collected (e.g. uncalculated formulas)
                            cell.value = sum_matrix.get((r, c), 0)
                        else:
                            pass # Text, empty, etc. Keep as is.

            # Save
            temp_dir = Path(settings.MEDIA_ROOT) / 'temp'
            temp_dir.mkdir(exist_ok=True)
            safe_filename = "".join(c for c in output_filename if c.isalnum() or c in (' ', '-', '_')).strip()
            output_path = temp_dir / f'{safe_filename}.xlsx'
            wb.save(output_path)
            wb.close()
            
            return FileResponse(
                open(output_path, 'rb'),
                as_attachment=True,
                filename=f'{safe_filename}.xlsx'
            )
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


# ============================================
# STYLE HELPERS
# ============================================

def copy_font(font):
    if not font: return None
    return Font(name=font.name, size=font.size, bold=font.bold, italic=font.italic, 
                vertAlign=font.vertAlign, underline=font.underline, strike=font.strike, color=font.color)

def copy_fill(fill):
    if not fill: return None
    return PatternFill(fill_type=fill.fill_type, start_color=fill.start_color, end_color=fill.end_color)

def copy_border(border):
    if not border: return None
    return Border(left=border.left, right=border.right, top=border.top, bottom=border.bottom, 
                  diagonal=border.diagonal, diagonal_direction=border.diagonal_direction, 
                  outline=border.outline, vertical=border.vertical, horizontal=border.horizontal)

def copy_alignment(alignment):
    if not alignment: return None
    return Alignment(horizontal=alignment.horizontal, vertical=alignment.vertical, 
                     text_rotation=alignment.text_rotation, wrap_text=alignment.wrap_text, 
                     shrink_to_fit=alignment.shrink_to_fit, indent=alignment.indent)
