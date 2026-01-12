"""
Excel File Consolidation for Django
Combines multiple Excel files by sheet, columns, and rows
"""
import os
import uuid
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings


@dataclass
class FileSource:
    """Source file configuration"""
    filepath: str
    responsible: str
    branch: str = ""
    cost_center: str = ""



def consolidate_files(
    files: List[Dict],
    sheet_name: str,
    start_column: str,
    end_column: str,
    start_row: int,
    end_row: int,
    group_by: str = "branch"
) -> str:
    """
    Consolidate multiple Excel files into one
    
    Args:
        files: List of dicts with: filepath, responsible, branch, cost_center
        sheet_name: Sheet to extract from (fallback)
        start_column: Start column letter
        end_column: End column letter
        start_row: Start row
        end_row: End row
        group_by: Field to group by
        
    Returns:
        Filename of generated file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidation"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Calculate column range
    start_col_idx = column_index_from_string(start_column.upper())
    end_col_idx = column_index_from_string(end_column.upper())
    num_cols = end_col_idx - start_col_idx + 1

    # Write headers
    current_row = 1
    headers = ["Source", "Feuille", "Responsable", "Branche", "Centre de Coût"]
    
    for i in range(num_cols):
        headers.append(f"Col {get_column_letter(start_col_idx + i)}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    current_row += 1

    # Group sources
    groups: Dict[str, List[Dict]] = {}
    for f in files:
        key = f.get(group_by, f.get("responsible", "Default"))
        if key not in groups:
            groups[key] = []
        groups[key].append(f)

    # Process each file
    for group_name, group_files in groups.items():
        # Group header
        group_cell = ws.cell(row=current_row, column=1, value=f"📁 {group_name}")
        group_cell.font = Font(bold=True, color="ED7D31")
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=len(headers))
        current_row += 1

        for file_info in group_files:
            filepath = file_info.get("filepath", "")
            
            if not filepath or not os.path.exists(filepath):
                continue

            try:
                source_wb = load_workbook(filepath, data_only=True)
                
                # Determine sheets to process
                sheets_to_process = file_info.get("sheets", [])
                if not sheets_to_process:
                    sheets_to_process = [sheet_name]

                for current_sheet_name in sheets_to_process:
                    if current_sheet_name not in source_wb.sheetnames:
                        continue

                    source_ws = source_wb[current_sheet_name]

                    for row in range(start_row, end_row + 1):
                        has_data = False
                        row_data = []
                        
                        for col in range(start_col_idx, end_col_idx + 1):
                            cell_value = source_ws.cell(row=row, column=col).value
                            row_data.append(cell_value)
                            if cell_value is not None:
                                has_data = True

                        if has_data:
                            ws.cell(row=current_row, column=1, value=os.path.basename(filepath))
                            ws.cell(row=current_row, column=2, value=current_sheet_name)
                            ws.cell(row=current_row, column=3, value=file_info.get("responsible", ""))
                            ws.cell(row=current_row, column=4, value=file_info.get("branch", ""))
                            ws.cell(row=current_row, column=5, value=file_info.get("cost_center", ""))

                            for col_offset, value in enumerate(row_data):
                                cell = ws.cell(row=current_row, column=6 + col_offset, value=value)
                                cell.border = border

                            current_row += 1

                source_wb.close()

            except Exception as e:
                print(f"Error processing {filepath}: {e}")
                continue

        current_row += 1

    # Auto-adjust columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save
    filename = f"Consolidation_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = Path(settings.MEDIA_ROOT) / "excel" / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)

    return filename


def analyze_with_ai(files: List[str], description: str, api_key: str) -> Dict:
    """
    Use AI to analyze files and suggest consolidation settings
    """
    import requests
    import re
    import json
    
    prompt = f"""Tu es un expert en consolidation de fichiers Excel.

Fichiers à consolider:
{chr(10).join(['- ' + f for f in files])}

Description utilisateur:
{description}

Génère une configuration JSON:
{{
    "sheet_name": "Nom feuille",
    "start_column": "E",
    "end_column": "P", 
    "start_row": 1,
    "end_row": 10,
    "group_by": "branch",
    "suggestions": "Notes..."
}}
"""
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    payload = {
        "model": "glm-4.7",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
        "max_tokens": 1024
    }
    
    response = requests.post(
        "https://open.bigmodel.cn/api/paas/v4/chat/completions",
        headers=headers,
        json=payload,
        timeout=60
    )
    
    if response.status_code != 200:
        raise ValueError(f"Erreur API: {response.status_code}")
    
    result = response.json()
    content = result["choices"][0]["message"]["content"]
    
    # Extract JSON
    if "```json" in content:
        match = re.search(r'```json\s*([\s\S]*?)\s*```', content)
        if match:
            content = match.group(1)
    
    start = content.find("{")
    if start == -1:
        raise ValueError("JSON invalide")
    
    depth = 0
    for i, char in enumerate(content[start:], start):
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return json.loads(content[start:i+1])
    
    raise ValueError("JSON incomplet")
